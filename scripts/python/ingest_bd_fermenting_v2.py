"""
ingest_bd_fermenting_v2.py — upload normalized FermentationData into bd_fermenting_v2.

Source: /var/www/maltytask/data/RawDB-normalized.xlsx
Sheets:
  FermentationData_DryHop    → event_type='DryHop'     ( ~279 rows)
  FermentationData_Reads     → event_type='Reads'       (~5262 rows)
  FermentationData_Purge     → event_type='Purge'       ( ~325 rows)
  FermentationData_ColdCrash → event_type='ColdCrash'   ( ~820 rows)

All four sheets fold into bd_fermenting_v2 via the event_type discriminator column.

OPERATOR SETUP: Before running, ensure xlsx is on VPS:
  scp /home/kluk/projects/maltytask/data/RawDB-normalized.xlsx \\
      maltyweb:/var/www/maltytask/data/RawDB-normalized.xlsx

Usage:
  python3 ingest_bd_fermenting_v2.py                     # dry-run all sheets
  python3 ingest_bd_fermenting_v2.py --apply             # live write
  python3 ingest_bd_fermenting_v2.py --sheet DryHop      # dry-run one sheet
  python3 ingest_bd_fermenting_v2.py --apply --verify    # live + verification
  python3 ingest_bd_fermenting_v2.py --limit 20          # dry-run first 20 rows

Sheets: DryHop | Reads | Purge | ColdCrash  (default: all)

Column map by sheet:
  DryHop:     timestamp(0) email(1) event_date(2) beer_raw(3) batch(4) recipe_id(5)
              category(6) line_idx(7) mi_id_resolved(8) raw_name(9) qty(10) unit(11)
              lot(12) confidence(13) parse_note(14) source_row(15)
  Reads:      Timestamp(0) Email(1) beer_raw(2) beer_recipe_id(3) batch(4)
              Gravity(5) pH(6) Temperature(7) Final_Comments(8)
  Purge:      Timestamp(0) Email(1) beer_raw(2) beer_recipe_id(3) batch(4)
              Comment_Purge(5) Final_Comments(6)
  ColdCrash:  Timestamp(0) Email(1) beer_raw(2) beer_recipe_id(3) batch(4)
              Comment_ColdCrash(5)
"""
from __future__ import annotations

import argparse
import hashlib
import json
import sys
from collections import Counter
from datetime import datetime, date
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

_SCRIPT_DIR = Path(__file__).parent
if str(_SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPT_DIR))

from lib_config import load as load_config
from lib_db import connect

# ── Constants ──────────────────────────────────────────────────────────────────

XLSX_PATH   = Path("/var/www/maltytask/data/RawDB-normalized.xlsx")
TARGET_TABLE = "bd_fermenting_v2"

FERMENTING_NK = frozenset({"submitted_at", "event_type", "beer_raw", "batch", "line_idx"})

_SHEETS = [
    ("FermentationData_DryHop",    "DryHop"),
    ("FermentationData_Reads",     "Reads"),
    ("FermentationData_Purge",     "Purge"),
    ("FermentationData_ColdCrash", "ColdCrash"),
]

# ── Type coercions (mirrors ingest_bd_brewing_v2.py) ──────────────────────────

def _s(v: Any) -> str | None:
    if v is None: return None
    s = str(v).strip()
    return s if s else None

def _i(v: Any) -> int | None:
    if v is None: return None
    if isinstance(v, bool): return int(v)
    if isinstance(v, (int, float)): return int(v)
    s = str(v).strip()
    if not s: return None
    try: return int(float(s))
    except (ValueError, TypeError): return None

def _f(v: Any) -> float | None:
    if v is None: return None
    if isinstance(v, bool): return float(v)
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip()
    if not s: return None
    try: return float(s)
    except (ValueError, TypeError): return None

def _dt(v: Any) -> str | None:
    if v is None: return None
    if isinstance(v, datetime): return v.strftime("%Y-%m-%d %H:%M:%S.%f")
    if isinstance(v, date): return datetime(v.year, v.month, v.day).strftime("%Y-%m-%d %H:%M:%S.000000")
    s = str(v).strip()
    if not s: return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try: return datetime.strptime(s, fmt).strftime("%Y-%m-%d %H:%M:%S.%f")
        except ValueError: pass
    return None

def _date(v: Any) -> str | None:
    if v is None: return None
    if isinstance(v, datetime): return v.strftime("%Y-%m-%d")
    if isinstance(v, date): return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s: return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y"):
        try: return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError: pass
    return None

def _row_hash(d: dict) -> str:
    return hashlib.sha256(
        json.dumps(d, sort_keys=True, default=str).encode("utf-8")
    ).hexdigest()

def _is_empty(row: tuple) -> bool:
    return not any(v is not None and str(v).strip() != "" for v in row)


# Meta columns excluded from the content hash.
_META_KEYS = frozenset({"row_hash", "is_tombstoned", "audit_flags",
                        "imported_at", "updated_at", "id"})


def _finalize(rows: list[dict]) -> list[dict]:
    """
    Two-step finalize for a single event_type's parsed rows:

      1. Reassign line_idx for non-DryHop collision groups. The early Reads data
         uses minute-precision timestamps, so multiple distinct readings on the
         same (submitted_at, beer_raw, batch) collide on the natural key when
         line_idx is fixed at 0. We give each row a deterministic within-group
         ordinal (content-sorted, so the assignment is stable across xlsx
         regenerations) — no reading is silently overwritten by the upsert.
         DryHop keeps its source line_idx (the ingredient position, already unique).

      2. Compute row_hash over the business columns (all columns minus _META_KEYS)
         AFTER line_idx is final, so the content guard reflects the stored row.
    """
    from collections import defaultdict
    groups: dict[tuple, list[dict]] = defaultdict(list)
    for row in rows:
        if row["event_type"] == "DryHop":
            continue
        key = (row["submitted_at"], row["beer_raw"], row["batch"])
        groups[key].append(row)

    for key, group in groups.items():
        if len(group) <= 1:
            continue  # no collision — line_idx 0 is fine
        group.sort(key=lambda r: (
            r["gravity"]     if r["gravity"]     is not None else float("-inf"),
            r["ph"]          if r["ph"]          is not None else float("-inf"),
            r["temperature"] if r["temperature"] is not None else float("-inf"),
            str(r["final_comments"]     or ""),
            str(r["comment_purge"]      or ""),
            str(r["comment_cold_crash"] or ""),
        ))
        for idx, row in enumerate(group):
            row["line_idx"] = idx

    for row in rows:
        canonical = {k: v for k, v in row.items() if k not in _META_KEYS}
        row["row_hash"] = _row_hash(canonical)
    return rows


# ── Lookup helpers ─────────────────────────────────────────────────────────────

def load_recipe_map(conn) -> dict[int, int]:
    with conn.cursor() as cur:
        cur.execute("SELECT id FROM ref_recipes")
        return {r["id"]: r["id"] for r in cur.fetchall()}

def load_mi_string_map(conn) -> dict[str, int]:
    with conn.cursor() as cur:
        cur.execute("SELECT id, mi_id FROM ref_mi WHERE mi_id IS NOT NULL")
        return {r["mi_id"]: int(r["id"]) for r in cur.fetchall()}


# ── UPSERT helper ──────────────────────────────────────────────────────────────

def upsert_rows(conn, rows: list[dict], dry_run: bool, label: str) -> None:
    if not rows:
        print(f"  [{label}] 0 rows — nothing to upsert")
        return

    cols = list(rows[0].keys())
    placeholders = ", ".join(["%s"] * len(cols))
    col_names = ", ".join(f"`{c}`" for c in cols)
    update_cols = [c for c in cols if c not in FERMENTING_NK
                   and c not in ("row_hash", "imported_at", "id")]
    update_clause = ", ".join(f"`{c}`=VALUES(`{c}`)" for c in update_cols) if update_cols else "row_hash=row_hash"

    sql = (
        f"INSERT INTO `{TARGET_TABLE}` ({col_names}) VALUES ({placeholders})\n"
        f"ON DUPLICATE KEY UPDATE {update_clause}"
    )

    if dry_run:
        print(f"  [{label}] DRY-RUN: would upsert {len(rows)} rows")
        for r in rows[:2]:
            preview = {k: str(v)[:40] for k, v in r.items() if v is not None}
            print(f"    sample: {json.dumps(preview)}")
        return

    batch_size = 500
    affected = 0
    with conn.cursor() as cur:
        for start in range(0, len(rows), batch_size):
            batch = rows[start:start + batch_size]
            values = [[r[c] for c in cols] for r in batch]
            cur.executemany(sql, values)
            affected += cur.rowcount
    conn.commit()
    print(f"  [{label}] upserted {len(rows)} rows (MySQL affected: {affected})")


# ── Per-sheet parsers ──────────────────────────────────────────────────────────

def parse_dryhop(ws, recipe_map: dict, mi_string_map: dict, limit: int | None) -> list[dict]:
    """
    FermentationData_DryHop: per-ingredient lines for dry-hop additions.
    Col map: timestamp(0) email(1) event_date(2) beer_raw(3) batch(4) recipe_id(5)
             category(6) line_idx(7) mi_id_resolved(8) raw_name(9) qty(10) unit(11)
             lot(12) confidence(13) parse_note(14) source_row(15)
    """
    rows_out: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if _is_empty(row): continue
        if limit is not None and len(rows_out) >= limit: break
        r = list(row) + [None] * max(0, 16 - len(row))

        submitted_at = _dt(r[0])
        if not submitted_at:
            continue  # submitted_at is part of NK — skip if absent

        beer_raw = _s(r[3]) or ""
        batch    = str(_i(r[4])) if r[4] is not None else ""
        line_idx = _i(r[7]) or 0

        recipe_id_raw = _i(r[5])
        audit_flags: list[str] = []

        recipe_id_fk: int | None = None
        if recipe_id_raw is not None and recipe_id_raw in recipe_map:
            recipe_id_fk = recipe_id_raw
        elif recipe_id_raw is not None:
            audit_flags.append(f"recipe_id_not_found:{recipe_id_raw}")

        mi_str = _s(r[8])
        mi_id_fk = mi_string_map.get(mi_str) if mi_str else None
        if mi_str and mi_id_fk is None:
            audit_flags.append(f"mi_unresolved:{mi_str}")

        unit_raw = _s(r[11])
        unit = unit_raw if unit_raw in ("kg", "g") else None

        canonical = {
            "submitted_at": submitted_at,
            "event_type":   "DryHop",
            "beer_raw":     beer_raw,
            "batch":        batch,
            "line_idx":     line_idx,
            "event_date":   _date(r[2]) or _date(r[0]),
            "email":        _s(r[1]),
            "recipe_id_fk": recipe_id_fk,
            "dh_category":  "hops_dry",
            "dh_mi_id_fk":  mi_id_fk,
            "dh_raw_name":  _s(r[9]),
            "dh_qty":       _f(r[10]),
            "dh_unit":      unit,
            "dh_lot":       _s(r[12]),
            "dh_confidence": _s(r[13]),
            "dh_parse_note": _s(r[14]),
            "dh_source_row": _i(r[15]),
        }
        rows_out.append({
            **canonical,
            "is_tombstoned": 0,
            "audit_flags":  ",".join(audit_flags) if audit_flags else None,
            # non-DryHop cols
            "gravity": None, "ph": None, "temperature": None,
            "comment_purge": None, "comment_cold_crash": None, "final_comments": None,
        })
    return rows_out


def parse_reads(ws, recipe_map: dict, limit: int | None) -> list[dict]:
    """
    FermentationData_Reads: gravity/pH/temperature readings per batch.
    Col map: Timestamp(0) Email(1) beer_raw(2) beer_recipe_id(3) batch(4)
             Gravity(5) pH(6) Temperature(7) Final_Comments(8)
    """
    rows_out: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if _is_empty(row): continue
        if limit is not None and len(rows_out) >= limit: break
        r = list(row) + [None] * max(0, 9 - len(row))

        submitted_at = _dt(r[0])
        if not submitted_at:
            continue

        beer_raw = _s(r[2]) or ""
        batch    = str(_i(r[4])) if r[4] is not None else ""

        recipe_id_raw = _i(r[3])
        audit_flags: list[str] = []

        recipe_id_fk: int | None = None
        if recipe_id_raw is not None and recipe_id_raw in recipe_map:
            recipe_id_fk = recipe_id_raw
        elif recipe_id_raw is not None:
            audit_flags.append(f"recipe_id_not_found:{recipe_id_raw}")

        canonical = {
            "submitted_at": submitted_at,
            "event_type":   "Reads",
            "beer_raw":     beer_raw,
            "batch":        batch,
            "line_idx":     0,
            "event_date":   _date(r[0]),
            "email":        _s(r[1]),
            "recipe_id_fk": recipe_id_fk,
            "gravity":      _f(r[5]),
            "ph":           _f(r[6]),
            "temperature":  _f(r[7]),
            "final_comments": _s(r[8]),
        }
        rows_out.append({
            **canonical,
            "is_tombstoned": 0,
            "audit_flags":  ",".join(audit_flags) if audit_flags else None,
            # non-Reads DryHop cols
            "dh_category": None, "dh_mi_id_fk": None, "dh_raw_name": None,
            "dh_qty": None, "dh_unit": None, "dh_lot": None,
            "dh_confidence": None, "dh_parse_note": None, "dh_source_row": None,
            # non-Reads other cols
            "comment_purge": None, "comment_cold_crash": None,
        })
    return rows_out


def parse_purge(ws, recipe_map: dict, limit: int | None) -> list[dict]:
    """
    FermentationData_Purge: CO2 purge events.
    Col map: Timestamp(0) Email(1) beer_raw(2) beer_recipe_id(3) batch(4)
             Comment_Purge(5) Final_Comments(6)
    """
    rows_out: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if _is_empty(row): continue
        if limit is not None and len(rows_out) >= limit: break
        r = list(row) + [None] * max(0, 7 - len(row))

        submitted_at = _dt(r[0])
        if not submitted_at:
            continue

        beer_raw = _s(r[2]) or ""
        batch    = str(_i(r[4])) if r[4] is not None else ""

        recipe_id_raw = _i(r[3])
        audit_flags: list[str] = []

        recipe_id_fk: int | None = None
        if recipe_id_raw is not None and recipe_id_raw in recipe_map:
            recipe_id_fk = recipe_id_raw
        elif recipe_id_raw is not None:
            audit_flags.append(f"recipe_id_not_found:{recipe_id_raw}")

        canonical = {
            "submitted_at":  submitted_at,
            "event_type":    "Purge",
            "beer_raw":      beer_raw,
            "batch":         batch,
            "line_idx":      0,
            "event_date":    _date(r[0]),
            "email":         _s(r[1]),
            "recipe_id_fk":  recipe_id_fk,
            "comment_purge": _s(r[5]),
            "final_comments": _s(r[6]),
        }
        rows_out.append({
            **canonical,
            "is_tombstoned": 0,
            "audit_flags":  ",".join(audit_flags) if audit_flags else None,
            # non-Purge cols
            "dh_category": None, "dh_mi_id_fk": None, "dh_raw_name": None,
            "dh_qty": None, "dh_unit": None, "dh_lot": None,
            "dh_confidence": None, "dh_parse_note": None, "dh_source_row": None,
            "gravity": None, "ph": None, "temperature": None,
            "comment_cold_crash": None,
        })
    return rows_out


def parse_coldcrash(ws, recipe_map: dict, limit: int | None) -> list[dict]:
    """
    FermentationData_ColdCrash: cold crash events.
    Col map: Timestamp(0) Email(1) beer_raw(2) beer_recipe_id(3) batch(4)
             Comment_ColdCrash(5)
    """
    rows_out: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if _is_empty(row): continue
        if limit is not None and len(rows_out) >= limit: break
        r = list(row) + [None] * max(0, 6 - len(row))

        submitted_at = _dt(r[0])
        if not submitted_at:
            continue

        beer_raw = _s(r[2]) or ""
        batch    = str(_i(r[4])) if r[4] is not None else ""

        recipe_id_raw = _i(r[3])
        audit_flags: list[str] = []

        recipe_id_fk: int | None = None
        if recipe_id_raw is not None and recipe_id_raw in recipe_map:
            recipe_id_fk = recipe_id_raw
        elif recipe_id_raw is not None:
            audit_flags.append(f"recipe_id_not_found:{recipe_id_raw}")

        canonical = {
            "submitted_at":      submitted_at,
            "event_type":        "ColdCrash",
            "beer_raw":          beer_raw,
            "batch":             batch,
            "line_idx":          0,
            "event_date":        _date(r[0]),
            "email":             _s(r[1]),
            "recipe_id_fk":      recipe_id_fk,
            "comment_cold_crash": _s(r[5]),
        }
        rows_out.append({
            **canonical,
            "is_tombstoned": 0,
            "audit_flags":  ",".join(audit_flags) if audit_flags else None,
            # non-ColdCrash cols
            "dh_category": None, "dh_mi_id_fk": None, "dh_raw_name": None,
            "dh_qty": None, "dh_unit": None, "dh_lot": None,
            "dh_confidence": None, "dh_parse_note": None, "dh_source_row": None,
            "gravity": None, "ph": None, "temperature": None,
            "comment_purge": None, "final_comments": None,
        })
    return rows_out


# ── Verification ───────────────────────────────────────────────────────────────

def run_verification(conn) -> None:
    checks = [
        ("total rows",
         "SELECT COUNT(*) AS n FROM bd_fermenting_v2"),
        ("per event_type",
         "SELECT event_type, COUNT(*) AS n FROM bd_fermenting_v2 GROUP BY event_type ORDER BY event_type"),
        ("recipe_id_fk NULL",
         "SELECT COUNT(*) AS n FROM bd_fermenting_v2 WHERE recipe_id_fk IS NULL"),
        ("DryHop mi_id_fk NULL",
         "SELECT COUNT(*) AS n FROM bd_fermenting_v2 WHERE event_type='DryHop' AND dh_mi_id_fk IS NULL"),
        ("rows with audit_flags",
         "SELECT COUNT(*) AS n FROM bd_fermenting_v2 WHERE audit_flags IS NOT NULL"),
    ]
    print("\n=== Verification ===")
    with conn.cursor() as cur:
        for label, sql in checks:
            cur.execute(sql)
            rows = cur.fetchall()
            print(f"\n  {label}:")
            for r in rows:
                print(f"    {dict(r)}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--apply",  action="store_true", help="Write to DB. Default dry-run.")
    parser.add_argument("--sheet",  choices=["DryHop", "Reads", "Purge", "ColdCrash"],
                        default=None, help="Process only this sheet (default: all).")
    parser.add_argument("--limit",  type=int, default=None, help="First N rows per sheet.")
    parser.add_argument("--verify", action="store_true", help="Run verification after upload.")
    args = parser.parse_args()

    dry_run = not args.apply
    sheets  = [args.sheet] if args.sheet else ["DryHop", "Reads", "Purge", "ColdCrash"]

    if not XLSX_PATH.exists():
        print(f"ERROR: xlsx not found at {XLSX_PATH}", file=sys.stderr)
        sys.exit(1)

    cfg  = load_config()
    conn = connect(cfg)

    recipe_map    = load_recipe_map(conn)
    mi_string_map = load_mi_string_map(conn)

    print(f"{'DRY-RUN' if dry_run else 'LIVE'} — sheets: {', '.join(sheets)}")
    print(f"Loaded: {len(recipe_map)} recipes, {len(mi_string_map)} MI string IDs\n")

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

    if "DryHop" in sheets:
        print("── DryHop ───────────────────────────────────────────────────────")
        rows = _finalize(parse_dryhop(wb["FermentationData_DryHop"], recipe_map, mi_string_map, args.limit))
        print(f"  parsed {len(rows)} rows")
        upsert_rows(conn, rows, dry_run, "DryHop")
        print()

    if "Reads" in sheets:
        print("── Reads ────────────────────────────────────────────────────────")
        rows = _finalize(parse_reads(wb["FermentationData_Reads"], recipe_map, args.limit))
        print(f"  parsed {len(rows)} rows")
        upsert_rows(conn, rows, dry_run, "Reads")
        print()

    if "Purge" in sheets:
        print("── Purge ────────────────────────────────────────────────────────")
        rows = _finalize(parse_purge(wb["FermentationData_Purge"], recipe_map, args.limit))
        print(f"  parsed {len(rows)} rows")
        upsert_rows(conn, rows, dry_run, "Purge")
        print()

    if "ColdCrash" in sheets:
        print("── ColdCrash ────────────────────────────────────────────────────")
        rows = _finalize(parse_coldcrash(wb["FermentationData_ColdCrash"], recipe_map, args.limit))
        print(f"  parsed {len(rows)} rows")
        upsert_rows(conn, rows, dry_run, "ColdCrash")
        print()

    wb.close()

    if not dry_run and args.verify:
        run_verification(conn)

    conn.close()
    print("Done.")


if __name__ == "__main__":
    main()
