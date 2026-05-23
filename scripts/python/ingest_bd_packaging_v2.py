"""
ingest_bd_packaging_v2.py — upload normalized PackagingData_Events into bd_packaging_v2.

Source: /var/www/maltytask/data/RawDB-normalized.xlsx, sheet PackagingData_Events.

OPERATOR SETUP: Before running, SCP the xlsx from the local maltytask machine:
  scp /home/kluk/projects/maltytask/data/RawDB-normalized.xlsx \\
      maltyweb:/var/www/maltytask/data/RawDB-normalized.xlsx

TODO (deferred): --full-sync tombstone mode — detect xlsx rows that disappeared vs
the previous ingest (deleted/edited form responses) and set is_tombstoned=1 on
orphaned DB rows. Current --apply run is append/upsert only; tombstoning requires
a stable stable-key comparison against a prior snapshot and is planned for Phase 1.B.

Column map (xlsx 0-based index → bd_packaging_v2 column):
  [0]  Timestamp                        → submitted_at
  [1]  Email Address                    → email
  [7]  Recettes Nébuleuse               → neb_beer
  [8]  Batch (neb)                      → neb_batch  (INT → VARCHAR)
  [9]  DLC (neb)                        → neb_dlc
  [10] Recettes Contract                → contract_beer
  [11] Batch (contract, raw)            → (used as fallback if col85 missing)
  [12] DLC (contract)                   → (unused — no target col)
  [52] Production Totale (main rows)    → prod_total_units
  [65] Quantité Unités (parallel rows)  → prod_total_units (for parallel) + special_qty_units
  [67] Commentaires                     → comments
  [69] Quel Client a été livré          → keg_client_delivered
  [81] nebuleuse_recipe_id (INT|None)   → used for recipe_id_fk (primary)
  [82] nebuleuse_format_suffix          → nebuleuse_format_suffix
  [84] contract_recipe_id (INT|None)    → used for recipe_id_fk (fallback)
  [85] contract_batch                   → contract_batch
  [86] contract_beer_raw                → used to populate contract_beer
  [87] row_origin                       → row_origin
  [88] run_type                         → run_type
  [89] unsaleable_units                 → unsaleable_units
  [90] qa_analyses_units                → qa_analyses_units
  [91] qa_library_units                 → qa_library_units
  [92] loss_liquid_other_units          → loss_liquid_other_units
  [93] loss_4pack_btl_units             → loss_4pack_btl_units
  [94] loss_4pack_can_units             → loss_4pack_can_units
  [95] loss_wrap_btl_units              → loss_wrap_btl_units
  [96] loss_wrap_can_units              → loss_wrap_can_units
  [97] loss_label_btl_units             → loss_label_btl_units
  [98] loss_keg_collar_units            → loss_keg_collar_units
  [99] loss_crown_cork_units            → loss_crown_cork_units
  [100] loss_can_lid_units              → loss_can_lid_units
  [101] loss_keg_save_units             → loss_keg_save_units
  [102] loss_container_btl_units        → loss_container_btl_units
  [103] loss_container_can_units        → loss_container_can_units
  [104] keg_client_delivered            → keg_client_delivered (normalized col from script)
  [105] new_liner_client (bool)         → new_liner_client (TINYINT)
  [106] new_liner_transport (bool)      → new_liner_transport (TINYINT)
  [107] is_white_label (bool)           → is_white_label (TINYINT)
  [108] white_label_name                → white_label_name
  [109] audit_flags                     → audit_flags
  [45]  Selection Canette_mi_id_resolved → (string mi_id → resolved to ref_mi.id)
  [49]  Selection Bouteille_mi_id_resolved → (string mi_id → resolved to ref_mi.id)

Usage:
  python3 ingest_bd_packaging_v2.py             # dry-run (default)
  python3 ingest_bd_packaging_v2.py --apply     # live write to bd_packaging_v2
  python3 ingest_bd_packaging_v2.py --limit 10  # dry-run first 10 rows only
"""
from __future__ import annotations

import argparse
import hashlib
import json
import sys
from collections import Counter
from datetime import datetime, date
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# Allow running from /var/www/maltytask (adds scripts/python to path for lib_* imports)
_SCRIPT_DIR = Path(__file__).parent
if str(_SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPT_DIR))

from lib_config import load as load_config
from lib_db import connect

# ── Constants ─────────────────────────────────────────────────────────────────

XLSX_PATH = Path("/var/www/maltytask/data/RawDB-normalized.xlsx")
SHEET_NAME = "PackagingData_Events"
TARGET_TABLE = "bd_packaging_v2"
SNAPSHOT_DIR = _SCRIPT_DIR / "data" / "bd-packaging-v2-snapshots"

# Natural-key columns (NEVER overwritten in ON DUPLICATE KEY UPDATE)
NATURAL_KEY_COLS = frozenset({
    "submitted_at", "neb_beer", "neb_batch",
    "contract_beer", "contract_batch",
    "row_origin", "nebuleuse_format_suffix",
})

# ── xlsx column indices (0-based) ─────────────────────────────────────────────
COL_TIMESTAMP        = 0
COL_EMAIL            = 1
COL_NEB_BEER         = 7
COL_NEB_BATCH        = 8     # raw int batch number for neb rows
COL_NEB_DLC          = 9
COL_CONTRACT_BEER    = 10
COL_CONTRACT_BATCH_RAW = 11  # raw batch for contract (fallback)
COL_PROD_TOTAL       = 52
COL_UNSALEABLE       = 53    # Invendable → maps to unsaleable_units (also at [89])
COL_QTY_UNITS        = 65
COL_COMMENTS         = 67
COL_KEG_CLIENT_RAW   = 69
COL_SEL_CAN_MI       = 45    # Selection Canette_mi_id_resolved (string mi_id)
COL_SEL_BOT_MI       = 49    # Selection Bouteille_mi_id_resolved (string mi_id)
COL_NEB_RECIPE_ID    = 81
COL_FORMAT_SUFFIX    = 82
COL_NEB_BEER_RAW     = 83    # normalized beer raw (same as col 7 but cleaned)
COL_CONTRACT_RECIPE_ID = 84
COL_CONTRACT_BATCH   = 85    # derived contract batch
COL_CONTRACT_BEER_RAW  = 86  # derived contract beer raw
COL_ROW_ORIGIN       = 87
COL_RUN_TYPE         = 88
COL_UNSALEABLE_UNITS = 89
COL_QA_ANALYSES      = 90
COL_QA_LIBRARY       = 91
COL_LOSS_LIQUID_OTHER = 92
COL_LOSS_4PACK_BTL   = 93
COL_LOSS_4PACK_CAN   = 94
COL_LOSS_WRAP_BTL    = 95
COL_LOSS_WRAP_CAN    = 96
COL_LOSS_LABEL_BTL   = 97
COL_LOSS_KEG_COLLAR  = 98
COL_LOSS_CROWN_CORK  = 99
COL_LOSS_CAN_LID     = 100
COL_LOSS_KEG_SAVE    = 101
COL_LOSS_CONTAINER_BTL = 102
COL_LOSS_CONTAINER_CAN = 103
COL_KEG_CLIENT       = 104   # normalized keg_client_delivered
COL_NEW_LINER_CLIENT = 105
COL_NEW_LINER_TRANSPORT = 106
COL_IS_WHITE_LABEL   = 107
COL_WHITE_LABEL_NAME = 108
COL_AUDIT_FLAGS      = 109

XLSX_WIDTH = 110


# ── Type coercions ────────────────────────────────────────────────────────────

def _s(v: Any) -> str | None:
    """Coerce to string or None."""
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _i(v: Any) -> int | None:
    """Coerce to int or None."""
    if v is None:
        return None
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    if not s:
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def _f(v: Any) -> float | None:
    """Coerce to float or None."""
    if v is None:
        return None
    if isinstance(v, bool):
        return float(v)
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _b(v: Any) -> int | None:
    """Coerce to TINYINT bool (0/1) or None."""
    if v is None:
        return None
    if isinstance(v, bool):
        return 1 if v else 0
    if isinstance(v, (int, float)):
        return 1 if v else 0
    s = str(v).strip().lower()
    if s in ("true", "1", "yes", "oui"):
        return 1
    if s in ("false", "0", "no", "non", ""):
        return 0
    return None


def _dt(v: Any) -> str | None:
    """Coerce to MySQL DATETIME(6) string or None."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S.%f")
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day).strftime("%Y-%m-%d %H:%M:%S.000000")
    s = str(v).strip()
    if not s:
        return None
    # Try ISO parse
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d %H:%M:%S.%f")
        except ValueError:
            pass
    return None


# ── Snapshot ──────────────────────────────────────────────────────────────────

def save_snapshot(conn) -> Path:
    """Write current bd_packaging_v2 state to snapshot JSON. Keep last 3."""
    SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.utcnow().strftime("%Y%m%dT%H%M%S")
    dest = SNAPSHOT_DIR / f"bd-packaging-v2-{ts}.json"

    with conn.cursor() as cur:
        cur.execute("SELECT COUNT(*) AS total FROM bd_packaging_v2")
        total = cur.fetchone()["total"]
        cur.execute(
            "SELECT audit_flags, COUNT(*) AS cnt FROM bd_packaging_v2"
            " GROUP BY audit_flags ORDER BY cnt DESC"
        )
        flag_dist = {(r["audit_flags"] or "NULL"): r["cnt"] for r in cur.fetchall()}

    snapshot = {
        "timestamp": ts,
        "table": TARGET_TABLE,
        "row_count": total,
        "audit_flags_distribution": flag_dist,
    }
    dest.write_text(json.dumps(snapshot, indent=2), encoding="utf-8")
    existing = sorted(SNAPSHOT_DIR.glob("bd-packaging-v2-*.json"))
    for old in existing[:-3]:
        old.unlink()
    return dest


# ── Lookup maps ───────────────────────────────────────────────────────────────

def load_sku_map(conn) -> dict[tuple[int, str], int]:
    """
    Pre-load (recipe_id, format_code) → sku_id map.
    Join: ref_recipes.sku_prefix || ref_packaging_formats.format_code = ref_skus.sku_code
    """
    sql = """
        SELECT rs.id AS sku_id, rs.recipe_id, rf.format_code
        FROM ref_skus rs
        JOIN ref_packaging_formats rf
          ON CONCAT(
               (SELECT rr.sku_prefix FROM ref_recipes rr WHERE rr.id = rs.recipe_id),
               rf.format_code
             ) = rs.sku_code
        WHERE rs.is_active = 1
    """
    with conn.cursor() as cur:
        cur.execute(sql)
        rows = cur.fetchall()
    m: dict[tuple[int, str], int] = {}
    for r in rows:
        if r["recipe_id"] is not None and r["format_code"] is not None:
            m[(int(r["recipe_id"]), str(r["format_code"]))] = int(r["sku_id"])
    return m


def load_mi_string_map(conn) -> dict[str, int]:
    """
    Load mi_id (string) → ref_mi.id (int) map for MI selection resolution.
    """
    with conn.cursor() as cur:
        cur.execute("SELECT id, mi_id FROM ref_mi WHERE mi_id IS NOT NULL")
        rows = cur.fetchall()
    return {r["mi_id"]: int(r["id"]) for r in rows}


# ── Row parsing ───────────────────────────────────────────────────────────────

def parse_row(
    raw: tuple,
    xlsx_row_num: int,
    sku_map: dict[tuple[int, str], int],
    mi_string_map: dict[str, int],
) -> dict:
    """
    Parse one xlsx data row into a bd_packaging_v2 insert dict.

    xlsx_row_num: 1-based row number in the sheet (row 1 = header, row 2 = first data).
    Returns a dict with all target columns + metadata (_audit_flags_extra, _row_num).
    """
    # Pad to expected width
    r: list = list(raw) + [None] * max(0, XLSX_WIDTH - len(raw))

    row_origin = _s(r[COL_ROW_ORIGIN]) or "main"
    audit_flags_src = _s(r[COL_AUDIT_FLAGS])
    extra_flags: list[str] = []

    # ── Recipe ID resolution ──────────────────────────────────────────────────
    neb_recipe_id  = _i(r[COL_NEB_RECIPE_ID])
    cont_recipe_id = _i(r[COL_CONTRACT_RECIPE_ID])
    # COALESCE: neb first, contract fallback
    recipe_id_fk = neb_recipe_id if neb_recipe_id is not None else cont_recipe_id
    if recipe_id_fk is None:
        extra_flags.append("recipe_unresolved")

    # ── Format suffix ─────────────────────────────────────────────────────────
    format_suffix = _s(r[COL_FORMAT_SUFFIX])

    # ── SKU resolution ────────────────────────────────────────────────────────
    sku_id_fk: int | None = None
    if recipe_id_fk is not None and format_suffix is not None:
        sku_id_fk = sku_map.get((recipe_id_fk, format_suffix))
        if sku_id_fk is None and row_origin == "main":
            extra_flags.append("sku_unresolved")
    elif row_origin == "main" and recipe_id_fk is not None:
        extra_flags.append("sku_unresolved")

    # ── Merge audit_flags ─────────────────────────────────────────────────────
    if extra_flags:
        if audit_flags_src:
            audit_flags = audit_flags_src + "," + ",".join(extra_flags)
        else:
            audit_flags = ",".join(extra_flags)
    else:
        audit_flags = audit_flags_src

    # ── Beer identity ─────────────────────────────────────────────────────────
    # neb_beer: prefer col83 (normalized raw) then col7 (original form field)
    neb_beer = _s(r[COL_NEB_BEER_RAW]) or _s(r[COL_NEB_BEER])
    # neb_batch: col8 (INT batch number for neb rows)
    neb_batch_val = _i(r[COL_NEB_BATCH])
    neb_batch = str(neb_batch_val) if neb_batch_val is not None else None
    neb_dlc = _dt(r[COL_NEB_DLC])

    # contract_beer: prefer col86 (derived) then col10 (original)
    contract_beer = _s(r[COL_CONTRACT_BEER_RAW]) or _s(r[COL_CONTRACT_BEER])
    # contract_batch: prefer col85 (derived) then col11 (raw)
    contract_batch_val = _s(r[COL_CONTRACT_BATCH])
    if contract_batch_val is None:
        cb_raw = _i(r[COL_CONTRACT_BATCH_RAW])
        contract_batch = str(cb_raw) if cb_raw is not None else None
    else:
        contract_batch = contract_batch_val

    # ── Production volumes ────────────────────────────────────────────────────
    # main rows: prod_total_units = col52 (Production Totale)
    # parallel rows: prod_total_units = col65 (Quantité Unités)
    if row_origin == "parallel":
        prod_total_units = _i(r[COL_QTY_UNITS])
    else:
        prod_total_units = _i(r[COL_PROD_TOTAL])
    special_qty_units = _i(r[COL_QTY_UNITS])  # always carry; meaningful for parallel

    # ── MI selections ─────────────────────────────────────────────────────────
    sel_can_str  = _s(r[COL_SEL_CAN_MI])
    sel_bot_str  = _s(r[COL_SEL_BOT_MI])
    selection_can_mi_id_fk  = mi_string_map.get(sel_can_str)  if sel_can_str  else None
    selection_bottle_mi_id_fk = mi_string_map.get(sel_bot_str) if sel_bot_str else None

    # ── Compute row_hash over canonical typed fields ───────────────────────────
    canonical = {
        "submitted_at":        _dt(r[COL_TIMESTAMP]),
        "email":               _s(r[COL_EMAIL]),
        "neb_beer":            neb_beer,
        "neb_batch":           neb_batch,
        "neb_dlc":             neb_dlc,
        "contract_beer":       contract_beer,
        "contract_batch":      contract_batch,
        "recipe_id_fk":        recipe_id_fk,
        "sku_id_fk":           sku_id_fk,
        "nebuleuse_format_suffix": format_suffix,
        "run_type":            _s(r[COL_RUN_TYPE]),
        "row_origin":          row_origin,
        "prod_total_units":    prod_total_units,
        "special_qty_units":   special_qty_units,
        "qa_analyses_units":   _i(r[COL_QA_ANALYSES]),
        "qa_library_units":    _i(r[COL_QA_LIBRARY]),
        "unsaleable_units":    _i(r[COL_UNSALEABLE_UNITS]),
        "loss_liquid_other_units": _f(r[COL_LOSS_LIQUID_OTHER]),
        "loss_4pack_btl_units":  _i(r[COL_LOSS_4PACK_BTL]),
        "loss_4pack_can_units":  _i(r[COL_LOSS_4PACK_CAN]),
        "loss_wrap_btl_units":   _i(r[COL_LOSS_WRAP_BTL]),
        "loss_wrap_can_units":   _i(r[COL_LOSS_WRAP_CAN]),
        "loss_label_btl_units":  _i(r[COL_LOSS_LABEL_BTL]),
        "loss_keg_collar_units": _i(r[COL_LOSS_KEG_COLLAR]),
        "loss_crown_cork_units": _i(r[COL_LOSS_CROWN_CORK]),
        "loss_can_lid_units":    _i(r[COL_LOSS_CAN_LID]),
        "loss_keg_save_units":   _i(r[COL_LOSS_KEG_SAVE]),
        "loss_container_btl_units": _i(r[COL_LOSS_CONTAINER_BTL]),
        "loss_container_can_units": _i(r[COL_LOSS_CONTAINER_CAN]),
        "keg_client_delivered":  _s(r[COL_KEG_CLIENT]),
        "new_liner_client":      _b(r[COL_NEW_LINER_CLIENT]),
        "new_liner_transport":   _b(r[COL_NEW_LINER_TRANSPORT]),
        "is_white_label":        _b(r[COL_IS_WHITE_LABEL]),
        "white_label_name":      _s(r[COL_WHITE_LABEL_NAME]),
        "audit_flags":           audit_flags,
        "comments":              _s(r[COL_COMMENTS]),
        "selection_can_mi_id_fk":    selection_can_mi_id_fk,
        "selection_bottle_mi_id_fk": selection_bottle_mi_id_fk,
    }
    rh = hashlib.sha256(
        json.dumps(canonical, sort_keys=True, default=str).encode("utf-8")
    ).hexdigest()

    return {
        # metadata
        "_row_num": xlsx_row_num,
        # identity
        "source_sheet_row_index": xlsx_row_num if row_origin == "main" else None,
        "row_origin":    row_origin,
        "row_hash":      rh,
        "is_tombstoned": 0,
        # timestamps
        "submitted_at":  canonical["submitted_at"],
        # beer identity
        "neb_beer":      neb_beer,
        "neb_batch":     neb_batch,
        "neb_dlc":       neb_dlc,
        "contract_beer": contract_beer,
        "contract_batch": contract_batch,
        # FK
        "recipe_id_fk":  recipe_id_fk,
        "sku_id_fk":     sku_id_fk,
        "client_fk":     None,   # not populated in this version
        # format + run type
        "nebuleuse_format_suffix": format_suffix,
        "run_type":      canonical["run_type"],
        # volumes
        "prod_total_units":  prod_total_units,
        "special_qty_units": special_qty_units,
        "vendable_hl":       None,  # not in xlsx — derived field, deferred
        # QA
        "qa_analyses_units": canonical["qa_analyses_units"],
        "qa_library_units":  canonical["qa_library_units"],
        # losses
        "unsaleable_units":      canonical["unsaleable_units"],
        "loss_liquid_other_units": canonical["loss_liquid_other_units"],
        "loss_4pack_btl_units":  canonical["loss_4pack_btl_units"],
        "loss_4pack_can_units":  canonical["loss_4pack_can_units"],
        "loss_wrap_btl_units":   canonical["loss_wrap_btl_units"],
        "loss_wrap_can_units":   canonical["loss_wrap_can_units"],
        "loss_label_btl_units":  canonical["loss_label_btl_units"],
        "loss_keg_collar_units": canonical["loss_keg_collar_units"],
        "loss_crown_cork_units": canonical["loss_crown_cork_units"],
        "loss_can_lid_units":    canonical["loss_can_lid_units"],
        "loss_keg_save_units":   canonical["loss_keg_save_units"],
        "loss_container_btl_units": canonical["loss_container_btl_units"],
        "loss_container_can_units": canonical["loss_container_can_units"],
        # keg-specific
        "keg_client_delivered":  canonical["keg_client_delivered"],
        "new_liner_client":      canonical["new_liner_client"],
        "new_liner_transport":   canonical["new_liner_transport"],
        # white label
        "is_white_label":    _b(r[COL_IS_WHITE_LABEL]) or 0,
        "white_label_name":  canonical["white_label_name"],
        # audit
        "audit_flags": audit_flags,
        "email":       canonical["email"],
        "comments":    canonical["comments"],
        # MI selections
        "selection_can_mi_id_fk":    selection_can_mi_id_fk,
        "selection_bottle_mi_id_fk": selection_bottle_mi_id_fk,
    }


# ── INSERT columns (subset of target — excludes id, imported_at, updated_at) ──

INSERT_COLS = [
    "source_sheet_row_index", "row_origin", "row_hash", "is_tombstoned",
    "submitted_at",
    "neb_beer", "neb_batch", "neb_dlc",
    "contract_beer", "contract_batch",
    "recipe_id_fk", "sku_id_fk", "client_fk",
    "nebuleuse_format_suffix", "run_type",
    "prod_total_units", "special_qty_units", "vendable_hl",
    "qa_analyses_units", "qa_library_units",
    "unsaleable_units", "loss_liquid_other_units",
    "loss_4pack_btl_units", "loss_4pack_can_units",
    "loss_wrap_btl_units", "loss_wrap_can_units",
    "loss_label_btl_units", "loss_keg_collar_units",
    "loss_crown_cork_units", "loss_can_lid_units",
    "loss_keg_save_units", "loss_container_btl_units", "loss_container_can_units",
    "keg_client_delivered", "new_liner_client", "new_liner_transport",
    "is_white_label", "white_label_name",
    "audit_flags", "email", "comments",
    "selection_can_mi_id_fk", "selection_bottle_mi_id_fk",
]

# UPDATE clause: all INSERT_COLS except natural-key and identity cols
_UPDATE_EXCLUDED = NATURAL_KEY_COLS | {"source_sheet_row_index", "row_origin", "is_tombstoned"}
UPDATE_COLS = [c for c in INSERT_COLS if c not in _UPDATE_EXCLUDED]


def _build_upsert_sql() -> str:
    cols_sql  = ", ".join(f"`{c}`" for c in INSERT_COLS)
    placeholders = ", ".join(["%s"] * len(INSERT_COLS))
    update_set = ",\n    ".join(f"`{c}` = VALUES(`{c}`)" for c in UPDATE_COLS)
    update_set += ",\n    `updated_at` = NOW()"
    return (
        f"INSERT INTO `{TARGET_TABLE}` ({cols_sql})\n"
        f"VALUES ({placeholders})\n"
        f"ON DUPLICATE KEY UPDATE\n"
        f"    {update_set}"
    )


UPSERT_SQL = _build_upsert_sql()

FAILURE_SQL = """
    INSERT INTO ingest_failures
        (run_id, source_tab, target_table, sheet_row_index, row_hash,
         reason_code, reason_text, raw_row)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
    ON DUPLICATE KEY UPDATE
        last_seen_at  = CURRENT_TIMESTAMP,
        reason_text   = VALUES(reason_text)
"""


def upsert_rows(
    conn, parsed: list[dict], *, dry_run: bool
) -> tuple[int, int, int, dict]:
    """
    INSERT ... ON DUPLICATE KEY UPDATE for each parsed row.
    Returns (inserted, updated_or_unchanged, failures, audit_flag_counts).
    """
    inserted = 0
    upserted = 0
    failures = 0
    flag_counts: Counter = Counter()

    for row in parsed:
        # tally audit flags
        flags_str = row.get("audit_flags") or ""
        for flag in flags_str.split(","):
            flag = flag.strip()
            if flag:
                flag_counts[flag] += 1

        if dry_run:
            continue

        values = [row[c] for c in INSERT_COLS]
        with conn.cursor() as cur:
            try:
                cur.execute(UPSERT_SQL, values)
                rc = cur.rowcount
                if rc == 1:
                    inserted += 1
                else:
                    upserted += 1
            except Exception as exc:
                failures += 1
                err_code = str(getattr(exc, "args", ["?"])[0])[:32]
                err_text = (str(exc.args[1]) if len(getattr(exc, "args", [])) > 1 else str(exc))[:512]
                try:
                    cur.execute(FAILURE_SQL, (
                        None,                          # run_id
                        SHEET_NAME,                    # source_tab
                        TARGET_TABLE,                  # target_table
                        row.get("source_sheet_row_index") or 0,
                        row.get("row_hash", ""),
                        err_code,
                        err_text,
                        json.dumps({k: v for k, v in row.items() if not k.startswith("_")},
                                   default=str),
                    ))
                except Exception as fe:
                    print(f"[WARNING] could not log failure: {fe}", file=sys.stderr)

    if not dry_run:
        conn.commit()

    return inserted, upserted, failures, dict(flag_counts)


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> int:
    ap = argparse.ArgumentParser(
        description="Upload normalized PackagingData_Events into bd_packaging_v2."
    )
    ap.add_argument("--apply", action="store_true",
                    help="Commit writes to DB (default is dry-run).")
    ap.add_argument("--limit", type=int, default=None, metavar="N",
                    help="Process only the first N data rows (debug).")
    args = ap.parse_args()
    dry_run = not args.apply

    # ── Verify xlsx exists ────────────────────────────────────────────────────
    if not XLSX_PATH.exists():
        print(
            f"ERROR: xlsx not found at {XLSX_PATH}\n"
            f"Run: scp /home/kluk/projects/maltytask/data/RawDB-normalized.xlsx "
            f"maltyweb:/var/www/maltytask/data/RawDB-normalized.xlsx",
            file=sys.stderr,
        )
        return 1

    prefix = "[DRY-RUN] " if dry_run else "[APPLY] "
    print(f"\n{prefix}Loading {XLSX_PATH} ...")

    # ── Open xlsx ─────────────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}", file=sys.stderr)
        return 1
    ws = wb[SHEET_NAME]

    # Read all rows (read_only returns generator — consume once)
    all_rows = list(ws.iter_rows(values_only=True))
    data_rows = all_rows[1:]   # skip header
    total_in_xlsx = len(data_rows)
    print(f"  xlsx rows (excl. header): {total_in_xlsx}")

    if args.limit:
        data_rows = data_rows[: args.limit]
        print(f"  limit applied: processing {len(data_rows)} rows")

    # ── Connect to DB ─────────────────────────────────────────────────────────
    cfg = load_config()
    conn = connect(cfg)

    try:
        # ── Load lookup maps ──────────────────────────────────────────────────
        print("  Loading lookup maps ...")
        sku_map      = load_sku_map(conn)
        mi_string_map = load_mi_string_map(conn)
        print(f"    SKU map entries:      {len(sku_map)}")
        print(f"    MI string map entries: {len(mi_string_map)}")

        # ── Parse rows ────────────────────────────────────────────────────────
        parsed: list[dict] = []
        for i, raw in enumerate(data_rows):
            xlsx_row_num = i + 2   # row 1 = header
            parsed.append(parse_row(raw, xlsx_row_num, sku_map, mi_string_map))

        # ── Stats ─────────────────────────────────────────────────────────────
        main_count     = sum(1 for r in parsed if r["row_origin"] == "main")
        parallel_count = sum(1 for r in parsed if r["row_origin"] == "parallel")
        null_recipe    = sum(1 for r in parsed if r["recipe_id_fk"] is None)
        null_sku_main  = sum(1 for r in parsed
                             if r["row_origin"] == "main" and r["sku_id_fk"] is None)

        run_type_dist = Counter(r["run_type"] for r in parsed)

        print(f"\n  Rows parsed:           {len(parsed)}")
        print(f"    main:                {main_count}")
        print(f"    parallel:            {parallel_count}")
        print(f"  Null recipe_id_fk:     {null_recipe}")
        print(f"  Null sku_id_fk (main): {null_sku_main}")
        print(f"  run_type distribution: {dict(run_type_dist)}")

        if dry_run:
            # Show what would happen
            with conn.cursor() as cur:
                cur.execute(f"SELECT COUNT(*) AS cnt FROM `{TARGET_TABLE}`")
                db_count = cur.fetchone()["cnt"]

            print(f"\n  Current DB rows:       {db_count}")
            print(f"  Would process:         {len(parsed)} rows (upsert)")

            # Tally audit flags
            flag_counts: Counter = Counter()
            for row in parsed:
                flags_str = row.get("audit_flags") or ""
                for flag in flags_str.split(","):
                    flag = flag.strip()
                    if flag:
                        flag_counts[flag] += 1

            print(f"\n  Audit flag distribution:")
            for flag, cnt in sorted(flag_counts.items(), key=lambda x: -x[1]):
                print(f"    {flag:<40s}: {cnt}")

            print(f"\n{prefix}No rows written. Re-run with --apply to commit.")
            return 0

        # ── Snapshot before write ─────────────────────────────────────────────
        snap_path = save_snapshot(conn)
        print(f"\n  Snapshot saved → {snap_path}")

        # ── Upsert ────────────────────────────────────────────────────────────
        print(f"  Inserting {len(parsed)} rows into {TARGET_TABLE} ...")
        inserted, upserted, failures, flag_counts = upsert_rows(
            conn, parsed, dry_run=False
        )

        print(f"\n{'=' * 52}")
        print(f"  Rows processed:     {len(parsed)}")
        print(f"  Inserts:            {inserted}")
        print(f"  Updates/unchanged:  {upserted}")
        print(f"  Tombstones:         0  (deferred — see TODO at top)")
        print(f"  Failures:           {failures}")
        print(f"\n  Audit flag distribution:")
        for flag, cnt in sorted(flag_counts.items(), key=lambda x: -x[1]):
            print(f"    {flag:<40s}: {cnt}")
        print(f"{'=' * 52}")

        if failures > 0:
            print(
                f"\n[WARNING] {failures} row(s) failed — see ingest_failures table.",
                file=sys.stderr,
            )

    finally:
        conn.close()

    return 0


if __name__ == "__main__":
    sys.exit(main())
