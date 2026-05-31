"""
append_mig236_to_xlsx.py — Patch the BATCH value on the 21 contract brew rows
in the xlsx BrewingData_Brewday sheet that currently have blank batch.

These 21 rows ALREADY EXIST in the xlsx with batch=None. Their blank batch is
why the loader's (beer, batch) natural key collapsed all same-beer blank rows
onto one survivor — the NK-collapse this whole backfill fixes. This script
writes the distinct batch number on each (matching mig 236), so a future
re-normalise reproduces the numbered rows instead of collapsing them.

BATCH-ONLY: recipe_id (col E) is intentionally left None to stay consistent
with every other row in the sheet (recipe_id_fk is resolved via a separate
path, not this column). It also corrects one genuine data-corruption in col L
(Pitched_From) for BLZ WestCoast batch 4 (2021-08-17) where the xlsx holds a
datetime object '2024-04-04' instead of the string '4,4'.

Target file: /home/kluk/projects/maltytask/data/RawDB-normalized.xlsx
(local copy; after patching, operator must scp it to the VPS:
  scp /home/kluk/projects/maltytask/data/RawDB-normalized.xlsx \\
      maltyweb:/var/www/maltytask/data/RawDB-normalized.xlsx
then re-run ingest_bd_brewing_v2.py --table brewday --apply to update
bd_brewing_brewday_v2 from the canonical xlsx source.)

Loader reads from: /var/www/maltytask/data/RawDB-normalized.xlsx (VPS only).
Local copy is the master edit source; VPS copy is the ingest input.

DO NOT RUN THIS SCRIPT YET — review the plan and patch preview first.

Usage (when approved):
  python3 scripts/python/append_mig236_to_xlsx.py

Safety:
  - Creates a backup at RawDB-normalized.xlsx.bak-mig236 before any write.
  - Prints a dry-run preview of all 21 patches before writing.
  - Writes dates as Python datetime/date objects (not strings) to avoid the
    known openpyxl DMY day/month swap bug.
  - Idempotent: re-running when batch/recipe_id are already set will report
    "already patched" and skip the write.
"""
from __future__ import annotations

import shutil
import sys
from datetime import datetime, date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ── Paths ──────────────────────────────────────────────────────────────────────
# Script lives in maltyweb/scripts/python/; xlsx is in maltytask/data/
_SCRIPT_DIR = Path(__file__).parent
# Navigate: maltyweb/scripts/python → maltyweb → sibling maltytask
_MALTYWEB_ROOT = _SCRIPT_DIR.parent.parent
_MALTYTASK_DATA = _MALTYWEB_ROOT.parent / "maltytask" / "data"
XLSX_PATH = _MALTYTASK_DATA / "RawDB-normalized.xlsx"
BACKUP_PATH = _MALTYTASK_DATA / "RawDB-normalized.xlsx.bak-mig236"

# ── Patch spec: (beer, event_date_str, batch_int, recipe_id_int, pitched_from_override) ──
# pitched_from_override is None unless the xlsx has a corrupt value to fix.
PATCHES: list[tuple[str, str, int, int, str | None]] = [
    ("BLZ Company - WestCoast Pale Ale", "2021-02-09", 1,  14, None),
    ("BLZ Company - WestCoast Pale Ale", "2021-04-29", 2,  14, None),
    ("BLZ Company - WestCoast Pale Ale", "2021-06-23", 3,  14, None),
    # Row 94: pitched_from in xlsx = datetime(2024,4,4) — corrupted; v1 DB = '4,4'
    ("BLZ Company - WestCoast Pale Ale", "2021-08-17", 4,  14, "4,4"),
    ("BLZ Company - WestCoast Pale Ale", "2021-09-08", 5,  14, None),
    ("BLZ Company - WestCoast Pale Ale", "2021-11-16", 6,  14, None),
    ("Brasserie du Château - Faya",       "2021-05-12", 1,  16, None),
    ("Brasserie du Château - Faya",       "2021-07-23", 2,  16, None),
    ("BLZ Company - Lager",               "2022-06-24", 1,  11, None),
    ("MeltingPote - Cropette",            "2021-07-26", 1,  41, None),
    ("Brasserie du Château - 4.4",        "2021-04-22", 3,  15, None),
    ("Brasserie du Château - 4.4",        "2021-06-11", 4,  15, None),
    ("Brasserie du Château - 4.4",        "2021-08-06", 5,  15, None),
    ("Brasserie du Château - 4.4",        "2021-11-15", 6,  15, None),
    ("Chien Bleu - Jasper",               "2021-05-21", 15, 21, None),
    ("Chien Bleu - Jasper",               "2021-09-13", 16, 21, None),
    ("Chien Bleu - Bamse",                "2021-07-06", 17, 20, None),
    ("Chien Bleu - Pomelo",               "2021-02-04", 14, 24, None),
    ("BadFish - Witshark",                "2021-01-08", 21,  9, None),
    ("BadFish - 915",                     "2021-02-02", 19,  7, None),
    ("BadFish - Cryo IPA",                "2021-04-27",  9,  8, None),
]

assert len(PATCHES) == 21, f"Expected 21 patches, got {len(PATCHES)}"

# Build lookup: (beer, event_date_str) → (batch, recipe_id, pitched_from_override)
patch_map: dict[tuple[str, str], tuple[int, int, str | None]] = {
    (beer, ev_date): (batch, recipe_id, pf_override)
    for beer, ev_date, batch, recipe_id, pf_override in PATCHES
}


def _cell_date_str(v) -> str | None:
    """Extract YYYY-MM-DD string from a cell value (datetime, date, or str)."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if len(s) >= 10:
        return s[:10]
    return None


def main(dry_run: bool = True) -> None:
    if not XLSX_PATH.exists():
        print(f"ERROR: xlsx not found at {XLSX_PATH}", file=sys.stderr)
        sys.exit(1)

    print(f"{'DRY-RUN' if dry_run else 'LIVE WRITE'}")
    print(f"Source: {XLSX_PATH}")
    if not dry_run:
        print(f"Backup: {BACKUP_PATH}")

    # Load workbook (NOT read_only — we need write access)
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["BrewingData_Brewday"]

    # Scan rows and build list of patches to apply
    found_map: dict[tuple[str, str], int] = {}   # (beer, date) → row_num
    to_patch: list[tuple[int, int, int, str | None]] = []  # (row_num, batch, recipe_id, pf_override)
    already_set: list[tuple[str, str, int]] = []  # (beer, date, row_num)

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        ts_cell   = row[0]
        beer_cell = row[2]
        batch_cell = row[3]

        ts_val   = ts_cell.value
        beer_val = beer_cell.value
        if ts_val is None or beer_val is None:
            continue

        ev_date = _cell_date_str(ts_val)
        beer    = str(beer_val).strip()
        key     = (beer, ev_date)

        if key not in patch_map:
            continue

        found_map[key] = row_num
        batch, recipe_id, pf_override = patch_map[key]
        current_batch = batch_cell.value

        if current_batch is not None:
            already_set.append((beer, ev_date, row_num))
            print(f"  [skip] Already has batch={current_batch!r}: {beer!r} {ev_date}")
            continue

        to_patch.append((row_num, batch, recipe_id, pf_override))
        pf_note = f" + pitched_from fix → {pf_override!r}" if pf_override else ""
        print(f"  [patch] row {row_num:4d}: {beer!r} {ev_date} → batch={batch} recipe_id={recipe_id}{pf_note}")

    # Check all 21 were found
    not_found = [k for k in patch_map if k not in found_map]
    if not_found:
        print(f"\nERROR: {len(not_found)} rows not found in xlsx:")
        for b, d in not_found:
            print(f"  {b!r} {d}")
        sys.exit(1)

    print(f"\n  Total: {len(to_patch)} rows to patch, {len(already_set)} already set")

    if dry_run:
        print("\nDry-run complete. To apply: run with dry_run=False or call main(dry_run=False)")
        wb.close()
        return

    if not to_patch:
        print("Nothing to patch.")
        wb.close()
        return

    # Backup first
    shutil.copy2(XLSX_PATH, BACKUP_PATH)
    print(f"\n  Backup written to {BACKUP_PATH}")

    # Re-open in write mode (non-read-only)
    wb2 = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws2 = wb2["BrewingData_Brewday"]

    # Scan again to get actual cell references (row_num is stable — no deletes)
    applied = 0
    for row_num_target, batch, recipe_id, pf_override in to_patch:
        # Write batch as integer (col D = index 3, 1-based col 4)
        ws2.cell(row=row_num_target, column=4).value = batch
        # NOTE: recipe_id (col E) is intentionally NOT written. Every other row in
        # this sheet — including the operator's numbered rows (8-12) — has
        # recipe_id=None; v2's recipe_id_fk is resolved via a separate path, not
        # this column. Writing it here would make these 21 the only populated rows
        # (inconsistent) and risk a future loader UPDATE clobbering. batch-only.
        # Fix corrupted pitched_from if needed (col L = index 11, 1-based col 12)
        if pf_override is not None:
            ws2.cell(row=row_num_target, column=12).value = pf_override
        applied += 1

    wb2.save(XLSX_PATH)
    wb2.close()
    print(f"  Applied {applied} patches. Saved to {XLSX_PATH}")
    print()
    print("Next steps:")
    print("  1. Verify the xlsx looks correct (re-run in dry_run mode → all 'already set')")
    print("  2. scp the xlsx to the VPS (keeps the ingest input in sync):")
    print(f"     scp {XLSX_PATH} maltyweb:/var/www/maltytask/data/RawDB-normalized.xlsx")
    print("  DO NOT auto-run the loader here. The DB is already correct via mig 236,")
    print("  and a full re-normalise has a pre-existing recipe_id-resolution concern")
    print("  (all rows, not just these) that must be cleared first — separate, deliberate act.")


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--apply", action="store_true",
                    help="Actually write to the xlsx (default: dry-run preview only)")
    args = ap.parse_args()
    main(dry_run=not args.apply)
