#!/usr/bin/env python3
"""
ingest_bc_sales_ledger.py — Incremental BC item-ledger → inv_sales_ledger connector.

Reads from the BC Online OData v4 web service
  Écritures_comptables_article_Excel
and upserts new entries into the canonical MySQL inv_sales_ledger table.

Incremental strategy: ANTI-JOIN on Entry_No (NOT a MAX watermark).
  BC does NOT assign Entry_No in strict chronological order — entries can be
  posted *after* a snapshot yet carry an Entry_No *below* the current MAX.  A
  `Entry_No gt MAX(bc_line_seq)` watermark would miss those "holes below the
  high-water mark" permanently.  Instead we:
    1. cutoff = MIN(posting_date) in DB  (scope floor — excludes out-of-scope
       pre-cutoff BC entries, e.g. pre-2021 Sale entries back to Entry_No 24).
    2. Pull all in-scope Sale Entry_Nos via a cheap $select=Entry_No projection
       (Posting_Date ge cutoff), server-paged via @odata.nextLink (NO $top).
    3. missing = scoped_api_entry_nos − db_bc_line_seqs.
    4. Fetch FULL rows for `missing` (OR-chained $filter, ~50/req — BC OData
       rejects the `in` operator with HTTP 501) and upsert idempotently.

Connection: pymysql direct to localhost:3306 via maltyweb lib_db pattern.
Credentials: /var/www/maltytask/config/db.env  (DB)
             /var/www/maltytask/config/bc.env   (BC OAuth2)

Idempotency:
  INSERT … ON DUPLICATE KEY UPDATE keyed on uq_bc_line_seq (bc_line_seq = Entry_No)
  AND on dedup_key GENERATED = CONCAT_WS('|', source_file, bc_line_seq).
  Both UNIQUE keys coalesce re-ingests to an UPDATE regardless of source_file.

Usage:
  # API incremental (default, dry-run):
  python3 scripts/python/ingest_bc_sales_ledger.py
  python3 scripts/python/ingest_bc_sales_ledger.py --apply
  python3 scripts/python/ingest_bc_sales_ledger.py --limit 50

  # xlsx one-off reload (fallback):
  python3 scripts/python/ingest_bc_sales_ledger.py --source xlsx --file PATH
  python3 scripts/python/ingest_bc_sales_ledger.py --source xlsx --file PATH --apply

  # Overlap verification (read-only, never writes — GATE A before arming cron):
  python3 scripts/python/ingest_bc_sales_ledger.py --verify-overlap 60000 61500
"""

from __future__ import annotations

import argparse
import csv
import os
import sys
from collections import Counter
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from urllib.parse import quote

# Allow running from /var/www/maltytask (adds scripts/python to path for lib_*)
_SCRIPT_DIR = Path(__file__).parent
if str(_SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPT_DIR))

import pymysql  # noqa: E402 — after sys.path fix

try:
    import requests  # noqa: E402
except ImportError:
    print(
        "ERROR: requests not installed. Run: pip install requests",
        file=sys.stderr,
    )
    sys.exit(1)

from lib_config import load as load_config  # noqa: E402

# ── Constants ──────────────────────────────────────────────────────────────────

# Source tag written to inv_sales_ledger.source_file for API-ingested rows.
API_SOURCE_FILE = "odata:Écritures_comptables_article_Excel"

# BC OData service (É URL-encoded as %C3%89 — Python urllib raises InvalidURL
# on raw non-ASCII in the path component).
_BC_SERVICE_NAME_ENCODED = "%C3%89critures_comptables_article_Excel"

BATCH_SIZE = 500

# ── Document-type map ─────────────────────────────────────────────────────────
# Maps BC Document_Type values → inv_sales_ledger.doc_type ENUM.
# English labels (OData API) and French labels (historical xlsx) both present.
DOC_TYPE_MAP: dict[str, str] = {
    # English (OData API)
    "Sales Shipment":        "shipment",
    "Sales Invoice":         "invoice",
    "Sales Credit Memo":     "credit",
    "Sales Return Receipt":  "return_receipt",
    # French (historical xlsx — kept for fallback; harmless)
    "Expédition vente":      "shipment",
    "Facture vente":         "invoice",
    "Avoir vente":           "credit",
    "Réception retour vente": "return_receipt",
}

# ── BC config loader ──────────────────────────────────────────────────────────

_BC_ENV_PATH = Path("/var/www/maltytask/config/bc.env")


def _load_bc_env() -> dict[str, str]:
    """
    Parse /var/www/maltytask/config/bc.env.
    Format: KEY=VALUE, one per line; blank lines and # comments ignored.
    Raises RuntimeError if the file is missing or a required key is absent.
    """
    path = _BC_ENV_PATH
    override = os.environ.get("MALTYTASK_BC_ENV")
    if override:
        path = Path(override)
    if not path.exists():
        raise RuntimeError(
            f"BC credentials not found at {path}.\n"
            f"Expected keys: BC_TENANT_ID, BC_CLIENT_ID, BC_CLIENT_SECRET, BC_ENVIRONMENT.\n"
            f"See: config/bc.env.example"
        )
    cfg: dict[str, str] = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            continue
        k, v = line.split("=", 1)
        cfg[k.strip()] = v.strip()
    for key in ("BC_TENANT_ID", "BC_CLIENT_ID", "BC_CLIENT_SECRET", "BC_ENVIRONMENT"):
        if key not in cfg:
            raise RuntimeError(f"Missing key in bc.env: {key}")
    return cfg


# ── OAuth2 token ──────────────────────────────────────────────────────────────

def _get_token(bc: dict[str, str]) -> str:
    """
    Fetch an OAuth2 client-credentials access token from Azure AD.
    Logs only HTTP status and expires_in — never logs the secret or the token.
    """
    tenant = bc["BC_TENANT_ID"]
    url = f"https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token"
    data = {
        "grant_type":    "client_credentials",
        "client_id":     bc["BC_CLIENT_ID"],
        "client_secret": bc["BC_CLIENT_SECRET"],
        "scope":         "https://api.businesscentral.dynamics.com/.default",
    }
    resp = requests.post(url, data=data, timeout=30)
    if resp.status_code != 200:
        raise RuntimeError(
            f"Token request failed: HTTP {resp.status_code} — "
            f"{resp.text[:200]}"
        )
    j = resp.json()
    expires_in = j.get("expires_in", "?")
    print(f"  [OAuth2] token acquired, expires_in={expires_in}s", flush=True)
    return j["access_token"]


# ── OData incremental fetch ───────────────────────────────────────────────────

def _odata_base(bc: dict[str, str]) -> str:
    """Return the OData service base URL (no query string)."""
    tenant = bc["BC_TENANT_ID"]
    env    = bc["BC_ENVIRONMENT"]
    return (
        f"https://api.businesscentral.dynamics.com/v2.0/{tenant}/{env}"
        f"/ODataV4/Company('NEBULEUSE')/{_BC_SERVICE_NAME_ENCODED}"
    )


# Max Entry_Nos per OR-chained $filter request (URL-length sane). BC OData on
# this service rejects the `in` operator (HTTP 501), so we OR-chain instead.
OR_CHAIN_BATCH = 50


def _build_scoped_projection_url(bc: dict[str, str], cutoff: str) -> str:
    """
    Build the cheap Entry_No projection URL for the in-scope Sale entries:
      Entry_Type eq 'Sale' and Posting_Date ge {cutoff}, $select=Entry_No.

    IMPORTANT: NO $top — a $top caps the result and suppresses @odata.nextLink,
    silently truncating. We rely on server paging via @odata.nextLink instead.

    Spaces → %20, the single quotes around Sale and the OData date literal are
    preserved by quote(safe="'_=").  The service name É is pre-encoded.
    """
    base = _odata_base(bc)
    # OData date literal for `ge` on an Edm.Date column is a bare ISO date
    # (no quotes): Posting_Date ge 2021-01-05
    filter_raw = f"Entry_Type eq 'Sale' and Posting_Date ge {cutoff}"
    filter_enc = quote(filter_raw, safe="'_=")
    select_enc = quote("Entry_No", safe="")
    return f"{base}?$filter={filter_enc}&$select={select_enc}"


def fetch_scoped_entry_nos(bc: dict[str, str], cutoff: str) -> set[int]:
    """
    Pull the full set of in-scope Sale Entry_Nos (Posting_Date ge cutoff) as a
    cheap $select=Entry_No projection.  Follows @odata.nextLink until exhausted
    (NO $top — see _build_scoped_projection_url).  Returns a set of Entry_Nos.

    The cutoff floor (= MIN(posting_date) in DB) auto-scopes to the period we
    already track and excludes pre-cutoff BC entries that are out of scope.

    Token hygiene: the bearer token is never logged.
    """
    token = _get_token(bc)
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept":        "application/json",
    }

    url: str | None = _build_scoped_projection_url(bc, cutoff)
    print(f"  [OData] scoped projection: Posting_Date ge {cutoff}, $select=Entry_No …",
          flush=True)

    entry_nos: set[int] = set()
    page = 0
    while url:
        resp = requests.get(url, headers=headers, timeout=120)
        if resp.status_code != 200:
            raise RuntimeError(
                f"OData projection request failed: HTTP {resp.status_code} — "
                f"{resp.text[:300]}"
            )
        j = resp.json()
        rows = j.get("value", [])
        for r in rows:
            en = r.get("Entry_No")
            if en is not None:
                entry_nos.add(int(en))
        page += 1
        print(
            f"  [OData] projection page {page}: {len(rows)} ids "
            f"(distinct total so far: {len(entry_nos):,})",
            flush=True,
        )
        url = j.get("@odata.nextLink")

    print(f"  [OData] projection complete: {len(entry_nos):,} in-scope Entry_Nos.",
          flush=True)
    return entry_nos


def fetch_odata_by_entry_nos(
    bc: dict[str, str],
    entry_nos: list[int],
) -> list[dict[str, Any]]:
    """
    Fetch FULL rows for an explicit set of Entry_Nos.  BC OData on this service
    does NOT support the `in` operator (HTTP 501), so we batch the Entry_Nos
    into OR-chained $filter requests (OR_CHAIN_BATCH per request, URL kept sane).
    Entry_Type eq 'Sale' is kept as a safety guard on every batch.

    Each batch is small enough to fit one page, but @odata.nextLink is still
    followed defensively in case a batch ever spans pages.

    Token hygiene: the bearer token is never logged.
    """
    if not entry_nos:
        return []

    token = _get_token(bc)
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept":        "application/json",
    }
    base = _odata_base(bc)
    order_enc = quote("Entry_No", safe="")

    sorted_seqs = sorted(entry_nos)
    all_rows: list[dict[str, Any]] = []
    n_batches = (len(sorted_seqs) + OR_CHAIN_BATCH - 1) // OR_CHAIN_BATCH

    for bi in range(0, len(sorted_seqs), OR_CHAIN_BATCH):
        batch = sorted_seqs[bi: bi + OR_CHAIN_BATCH]
        or_chain = " or ".join(f"Entry_No eq {s}" for s in batch)
        filter_raw = f"({or_chain}) and Entry_Type eq 'Sale'"
        filter_enc = quote(filter_raw, safe="'_=()")
        url: str | None = f"{base}?$filter={filter_enc}&$orderby={order_enc}"
        batch_no = bi // OR_CHAIN_BATCH + 1

        while url:
            resp = requests.get(url, headers=headers, timeout=60)
            if resp.status_code != 200:
                raise RuntimeError(
                    f"OData by-Entry_No request failed (batch {batch_no}): "
                    f"HTTP {resp.status_code} — {resp.text[:300]}"
                )
            j = resp.json()
            rows = j.get("value", [])
            all_rows.extend(rows)
            url = j.get("@odata.nextLink")

        print(
            f"  [OData] full-row batch {batch_no}/{n_batches}: "
            f"requested {len(batch)} ids (total rows so far: {len(all_rows)})",
            flush=True,
        )

    print(f"  [OData] full-row fetch complete: {len(all_rows)} rows.", flush=True)
    return all_rows


def fetch_odata_range(
    bc: dict[str, str],
    from_seq: int,
    to_seq: int,
) -> list[dict[str, Any]]:
    """
    Fetch BC item-ledger entries with Entry_No in [from_seq, to_seq] (inclusive)
    and Entry_Type = 'Sale'.  Used by --verify-overlap (read-only).
    """
    token = _get_token(bc)
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept":        "application/json",
    }

    tenant = bc["BC_TENANT_ID"]
    env    = bc["BC_ENVIRONMENT"]
    base   = (
        f"https://api.businesscentral.dynamics.com/v2.0/{tenant}/{env}"
        f"/ODataV4/Company('NEBULEUSE')/{_BC_SERVICE_NAME_ENCODED}"
    )
    filter_raw = (
        f"Entry_No ge {from_seq} and Entry_No le {to_seq} "
        f"and Entry_Type eq 'Sale'"
    )
    filter_enc = quote(filter_raw, safe="'_=")
    order_enc  = quote("Entry_No", safe="")
    url = f"{base}?$filter={filter_enc}&$orderby={order_enc}"

    all_rows: list[dict[str, Any]] = []
    page = 0
    while url:
        resp = requests.get(url, headers=headers, timeout=60)
        if resp.status_code != 200:
            raise RuntimeError(
                f"OData range request failed: HTTP {resp.status_code} — "
                f"{resp.text[:300]}"
            )
        j = resp.json()
        rows = j.get("value", [])
        all_rows.extend(rows)
        page += 1
        print(
            f"  [OData] range page {page}: {len(rows)} rows "
            f"(total so far: {len(all_rows)})",
            flush=True,
        )
        url = j.get("@odata.nextLink")

    return all_rows


# ── Reference-data loaders ─────────────────────────────────────────────────────

def load_customers(conn: pymysql.connections.Connection) -> dict[str, int]:
    """
    Returns {bc_customer_no: ref_customers.id} for all customers.
    Inactive rows included — historic ledger entries may reference retired customers.
    """
    with conn.cursor() as c:
        c.execute(
            "SELECT bc_customer_no, id FROM ref_customers "
            "WHERE bc_customer_no IS NOT NULL"
        )
        return {str(r["bc_customer_no"]).strip(): int(r["id"]) for r in c.fetchall()}


def load_skus(
    conn: pymysql.connections.Connection,
) -> tuple[dict[str, tuple[int, float | None]], dict[str, tuple[int, float | None]]]:
    """
    Returns:
      active_skus : {sku_code: (id, hl_per_unit_or_None)}  — active ref_skus
      alias_map   : {alias:   (canonical_sku_id, hl_per_unit_or_None)}
    Resolution order: direct match → alias.
    Inactive SKUs excluded from direct map — they should come via ref_sku_aliases.
    """
    with conn.cursor() as c:
        c.execute(
            "SELECT sku_code, id, hl_per_unit FROM ref_skus WHERE is_active = 1"
        )
        active_skus: dict[str, tuple[int, float | None]] = {
            str(r["sku_code"]).strip(): (
                int(r["id"]),
                float(r["hl_per_unit"]) if r["hl_per_unit"] is not None else None,
            )
            for r in c.fetchall()
        }

        c.execute(
            """
            SELECT ra.alias, ra.canonical_sku_id, rs.hl_per_unit
              FROM ref_sku_aliases ra
              JOIN ref_skus rs ON rs.id = ra.canonical_sku_id
            """
        )
        alias_map: dict[str, tuple[int, float | None]] = {
            str(r["alias"]).strip(): (
                int(r["canonical_sku_id"]),
                float(r["hl_per_unit"]) if r["hl_per_unit"] is not None else None,
            )
            for r in c.fetchall()
        }

    return active_skus, alias_map


def get_scope_cutoff(conn: pymysql.connections.Connection) -> str | None:
    """
    Returns MIN(posting_date) from inv_sales_ledger as an ISO 'YYYY-MM-DD'
    string — the scope floor for the incremental anti-join.  This auto-scopes
    the projection to the period we already track and excludes pre-cutoff BC
    entries (e.g. pre-2021 Sale entries back to Entry_No 24) that are out of
    scope and must NOT be loaded.

    Returns None if the table is empty (first-ever load should use --source xlsx).
    """
    with conn.cursor() as c:
        c.execute("SELECT MIN(posting_date) AS cutoff FROM inv_sales_ledger")
        row = c.fetchone()
        cutoff = row["cutoff"] if row else None
        if cutoff is None:
            return None
        # pymysql returns a datetime.date for a DATE column
        return cutoff.isoformat() if hasattr(cutoff, "isoformat") else str(cutoff)


def load_db_entry_nos(conn: pymysql.connections.Connection) -> set[int]:
    """
    Returns the full set of bc_line_seq (= BC Entry_No) already in
    inv_sales_ledger.  Used as the left side of the incremental anti-join.
    """
    with conn.cursor() as c:
        c.execute(
            "SELECT bc_line_seq FROM inv_sales_ledger WHERE bc_line_seq IS NOT NULL"
        )
        return {int(r["bc_line_seq"]) for r in c.fetchall()}


# ── Number parser ──────────────────────────────────────────────────────────────

def parse_swiss_number(v: Any, default: Decimal | None = None) -> Decimal | None:
    """
    Safe numeric parser for BC cells (xlsx or OData string/int/float).
    Handles int/float (native), str with trailing dot ("-1."), str with
    thousands separator (",").  Returns Decimal for precision, or default.
    """
    if v is None:
        return default
    if isinstance(v, (int, float)):
        try:
            return Decimal(str(v))
        except InvalidOperation:
            return default
    s = str(v).strip()
    if s in ("", "-"):
        return default
    s = s.replace(",", "")
    # Strip trailing dot (some BC exports have "-1.")
    if s.endswith("."):
        s = s[:-1]
    try:
        return Decimal(s)
    except InvalidOperation:
        return default


# ── API row parser ─────────────────────────────────────────────────────────────

def parse_api_row(
    api_dict: dict[str, Any],
    customers: dict[str, int],
    active_skus: dict[str, tuple[int, float | None]],
    alias_map: dict[str, tuple[int, float | None]],
    unmapped_doc_types: Counter,
) -> dict[str, Any] | None:
    """
    Map one BC OData API dict → the same row dict that INSERT_SQL expects.
    Returns None if Item_No is empty (skip silently — not a sale line).

    API field → DB column mapping:
      Entry_No            → bc_line_seq
      Posting_Date        → posting_date  (ISO 'YYYY-MM-DD' string from API)
      Document_Type       → doc_type (via DOC_TYPE_MAP)
      Document_No         → bc_document_no
      Item_No             → sku_code_raw
      Quantity            → qty_signed
      Invoiced_Quantity   → qty_invoiced
      Sales_Amount_Actual → sales_amount_chf
      Source_No           → bc_source_no  (customer number)
    """
    # Item_No — NOT NULL in DDL; skip if empty
    item_no = api_dict.get("Item_No")
    sku_code_raw = str(item_no).strip() if item_no else ""
    if not sku_code_raw:
        return None

    # Posting_Date — API returns ISO date string 'YYYY-MM-DD'
    pd_raw = api_dict.get("Posting_Date")
    if not pd_raw:
        raise ValueError(f"Missing Posting_Date in Entry_No={api_dict.get('Entry_No')}")
    # Accept both 'YYYY-MM-DD' and 'YYYY-MM-DDThh:mm:ss'
    pd_str = str(pd_raw).strip()
    if "T" in pd_str:
        pd_str = pd_str.split("T")[0]
    posting_date = pd_str  # stored as ISO string — MySQL DATE accepts 'YYYY-MM-DD'

    # bc_line_seq (Entry_No)
    entry_no = api_dict.get("Entry_No")
    bc_line_seq = int(entry_no) if entry_no is not None else None

    # Document_Type → mapped ENUM
    doc_type_raw_val = api_dict.get("Document_Type")
    doc_type_raw = str(doc_type_raw_val).strip() if doc_type_raw_val is not None else ""
    doc_type = DOC_TYPE_MAP.get(doc_type_raw)
    if doc_type is None:
        unmapped_doc_types[doc_type_raw] += 1
        doc_type = "shipment"  # visible in report; --apply aborts if any unmapped

    bc_document_no = str(api_dict["Document_No"]).strip() if api_dict.get("Document_No") else None

    bc_source_no_val = api_dict.get("Source_No")
    bc_source_no = str(bc_source_no_val).strip() if bc_source_no_val else None

    qty_signed    = parse_swiss_number(api_dict.get("Quantity"),            Decimal("0"))
    if qty_signed is None:
        qty_signed = Decimal("0")
    qty_invoiced      = parse_swiss_number(api_dict.get("Invoiced_Quantity"))
    sales_amount_chf  = parse_swiss_number(api_dict.get("Sales_Amount_Actual"))

    # Customer resolution — refuse-don't-NULL discipline: unresolved → NULL FK
    customer_id_fk: int | None = None
    if bc_source_no:
        customer_id_fk = customers.get(bc_source_no)

    # SKU resolution: direct → alias; unresolved → NULL FK (never guess)
    sku_id_fk: int | None = None
    hl_per_unit: float | None = None
    direct = active_skus.get(sku_code_raw)
    if direct is not None:
        sku_id_fk, hl_per_unit = direct
    else:
        aliased = alias_map.get(sku_code_raw)
        if aliased is not None:
            sku_id_fk, hl_per_unit = aliased

    hl_resolved: Decimal | None = None
    if sku_id_fk is not None and hl_per_unit is not None:
        hl_resolved = qty_signed * Decimal(str(hl_per_unit))

    return {
        "posting_date":     posting_date,
        "doc_type":         doc_type,
        "doc_type_raw":     doc_type_raw or None,
        "bc_document_no":   bc_document_no,
        "bc_line_seq":      bc_line_seq,
        "bc_source_no":     bc_source_no,
        "customer_id_fk":   customer_id_fk,
        "sku_code_raw":     sku_code_raw,
        "sku_id_fk":        sku_id_fk,
        "qty_signed":       qty_signed,
        "qty_invoiced":     qty_invoiced,
        "sales_amount_chf": sales_amount_chf,
        "hl_resolved":      hl_resolved,
        "source_file":      API_SOURCE_FILE,
    }


# ── xlsx loader (fallback path) ────────────────────────────────────────────────

# xlsx import deferred to here to avoid hard dependency when using --source api
def _import_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        print(
            "ERROR: openpyxl not installed. Run: pip install openpyxl",
            file=sys.stderr,
        )
        sys.exit(1)


XLSX_SHEET_NAME = "Écritures comptables article"

EXPECTED_XLSX_HEADERS = {
    "Date comptabilisation",
    "Type document",
    "N° document",
    "N° article",
    "Quantité",
    "Quantité facturée",
    "Montant vente (réel)",
    "N° séquence",
    "N° origine",
}


def _build_col_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    col_map = {str(h).strip(): i for i, h in enumerate(header_row) if h is not None}
    missing = EXPECTED_XLSX_HEADERS - set(col_map)
    if missing:
        raise ValueError(
            f"Missing expected xlsx column(s): {sorted(missing)}\n"
            f"Found: {sorted(col_map)}"
        )
    return col_map


def _parse_xlsx_row(
    raw: tuple[Any, ...],
    col: dict[str, int],
    customers: dict[str, int],
    active_skus: dict[str, tuple[int, float | None]],
    alias_map: dict[str, tuple[int, float | None]],
    source_file: str,
    unmapped_doc_types: Counter,
) -> dict[str, Any] | None:
    """Parse one xlsx data row. Returns None if N° article is empty."""
    sku_raw = raw[col["N° article"]]
    sku_code_raw = str(sku_raw).strip() if sku_raw is not None else ""
    if not sku_code_raw:
        return None

    d = raw[col["Date comptabilisation"]]
    if isinstance(d, datetime):
        posting_date = str(d.date())
    elif isinstance(d, str) and d.strip():
        for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
            try:
                posting_date = datetime.strptime(d.strip(), fmt).strftime("%Y-%m-%d")
                break
            except ValueError:
                continue
        else:
            raise ValueError(f"Unparseable date: {d!r}")
    else:
        raise ValueError("Missing posting_date in xlsx row")

    doc_type_raw_val = raw[col["Type document"]]
    doc_type_raw = str(doc_type_raw_val).strip() if doc_type_raw_val is not None else ""
    doc_type = DOC_TYPE_MAP.get(doc_type_raw)
    if doc_type is None:
        unmapped_doc_types[doc_type_raw] += 1
        doc_type = "shipment"

    bc_document_no = str(raw[col["N° document"]]).strip() if raw[col["N° document"]] else None

    bc_line_seq_val = raw[col["N° séquence"]]
    bc_line_seq = int(bc_line_seq_val) if bc_line_seq_val is not None else None

    bc_source_no_val = raw[col["N° origine"]]
    bc_source_no = str(bc_source_no_val).strip() if bc_source_no_val is not None else None

    qty_signed = parse_swiss_number(raw[col["Quantité"]], Decimal("0"))
    if qty_signed is None:
        qty_signed = Decimal("0")
    qty_invoiced      = parse_swiss_number(raw[col["Quantité facturée"]])
    sales_amount_chf  = parse_swiss_number(raw[col["Montant vente (réel)"]])

    customer_id_fk: int | None = None
    if bc_source_no:
        customer_id_fk = customers.get(bc_source_no)

    sku_id_fk: int | None = None
    hl_per_unit: float | None = None
    direct = active_skus.get(sku_code_raw)
    if direct is not None:
        sku_id_fk, hl_per_unit = direct
    else:
        aliased = alias_map.get(sku_code_raw)
        if aliased is not None:
            sku_id_fk, hl_per_unit = aliased

    hl_resolved: Decimal | None = None
    if sku_id_fk is not None and hl_per_unit is not None:
        hl_resolved = qty_signed * Decimal(str(hl_per_unit))

    return {
        "posting_date":     posting_date,
        "doc_type":         doc_type,
        "doc_type_raw":     doc_type_raw or None,
        "bc_document_no":   bc_document_no,
        "bc_line_seq":      bc_line_seq,
        "bc_source_no":     bc_source_no,
        "customer_id_fk":   customer_id_fk,
        "sku_code_raw":     sku_code_raw,
        "sku_id_fk":        sku_id_fk,
        "qty_signed":       qty_signed,
        "qty_invoiced":     qty_invoiced,
        "sales_amount_chf": sales_amount_chf,
        "hl_resolved":      hl_resolved,
        "source_file":      source_file,
    }


def parse_xlsx(
    path: Path,
    customers: dict[str, int],
    active_skus: dict[str, tuple[int, float | None]],
    alias_map: dict[str, tuple[int, float | None]],
    limit: int | None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Parse BC ledger xlsx, return (rows, stats)."""
    openpyxl = _import_openpyxl()
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    if XLSX_SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Sheet {XLSX_SHEET_NAME!r} not found. Available: {wb.sheetnames}"
        )
    ws = wb[XLSX_SHEET_NAME]
    source_file = path.name

    rows_parsed: list[dict[str, Any]] = []
    skipped_no_article = 0
    unmapped_doc_types: Counter = Counter()
    col: dict[str, int] = {}

    for i, raw in enumerate(ws.iter_rows(min_row=1, values_only=True)):
        if i == 0:
            col = _build_col_map(raw)
            continue
        parsed = _parse_xlsx_row(
            raw, col, customers, active_skus, alias_map,
            source_file, unmapped_doc_types,
        )
        if parsed is None:
            skipped_no_article += 1
            continue
        rows_parsed.append(parsed)
        if limit and len(rows_parsed) >= limit:
            break

    wb.close()
    stats = {
        "source":                source_file,
        "rows_data":             len(rows_parsed),
        "rows_skipped_no_art":   skipped_no_article,
        "unmapped_doc_types":    dict(unmapped_doc_types),
    }
    return rows_parsed, stats


# ── DB writer ──────────────────────────────────────────────────────────────────

INSERT_SQL = """
INSERT INTO inv_sales_ledger
  (posting_date, doc_type, doc_type_raw, bc_document_no, bc_line_seq,
   bc_source_no, customer_id_fk, sku_code_raw, sku_id_fk,
   qty_signed, qty_invoiced, sales_amount_chf, hl_resolved, source_file)
VALUES
  (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
ON DUPLICATE KEY UPDATE
  customer_id_fk   = VALUES(customer_id_fk),
  sku_id_fk        = VALUES(sku_id_fk),
  hl_resolved      = VALUES(hl_resolved),
  posting_date     = VALUES(posting_date),
  doc_type         = VALUES(doc_type),
  doc_type_raw     = VALUES(doc_type_raw),
  bc_source_no     = VALUES(bc_source_no),
  qty_signed       = VALUES(qty_signed),
  qty_invoiced     = VALUES(qty_invoiced),
  sales_amount_chf = VALUES(sales_amount_chf)
"""


def insert_rows(
    conn: pymysql.connections.Connection,
    rows: list[dict[str, Any]],
) -> int:
    """
    Upsert rows in batches.
    Returns MySQL aggregate rowcount signal (1=new, 2=updated, 0=unchanged per row).
    Both uq_bc_line_seq and uniq_sales_ledger_dedup can trigger the UPDATE path.
    """
    total_rc = 0
    with conn.cursor() as c:
        for start in range(0, len(rows), BATCH_SIZE):
            batch = rows[start: start + BATCH_SIZE]
            params = [
                (
                    r["posting_date"],
                    r["doc_type"],
                    r["doc_type_raw"],
                    r["bc_document_no"],
                    r["bc_line_seq"],
                    r["bc_source_no"],
                    r["customer_id_fk"],
                    r["sku_code_raw"],
                    r["sku_id_fk"],
                    str(r["qty_signed"]),
                    str(r["qty_invoiced"])     if r["qty_invoiced"]     is not None else None,
                    str(r["sales_amount_chf"]) if r["sales_amount_chf"] is not None else None,
                    str(r["hl_resolved"])      if r["hl_resolved"]      is not None else None,
                    r["source_file"],
                )
                for r in batch
            ]
            c.executemany(INSERT_SQL, params)
            total_rc += c.rowcount
        conn.commit()
    return total_rc


# ── Dry-run report ─────────────────────────────────────────────────────────────

def build_dry_run_report(
    rows: list[dict[str, Any]],
    stats: dict[str, Any],
) -> None:
    """Print the full dry-run report to stdout."""
    n = len(rows)
    print()
    print("=" * 70)
    print("DRY-RUN REPORT — ingest_bc_sales_ledger.py")
    print("=" * 70)

    if stats.get("scope_cutoff") is not None:
        print(f"\n   Incremental strategy             : ANTI-JOIN on Entry_No")
        print(f"   Scope cutoff (MIN posting_date)  : {stats['scope_cutoff']}")
        if "missing_count" in stats:
            print(f"   Missing Entry_Nos (to insert)    : {stats['missing_count']:,}")
        if rows:
            min_seq = min(r["bc_line_seq"] for r in rows if r["bc_line_seq"] is not None)
            max_seq = max(r["bc_line_seq"] for r in rows if r["bc_line_seq"] is not None)
            print(f"   Entry_No range in to-insert set  : {min_seq:,} – {max_seq:,}")

    print()
    print("1. PARSED ROW COUNT + DATE SPAN")
    print(f"   Total rows parsed             : {n:,}")
    if "rows_skipped_no_art" in stats:
        print(f"   Rows skipped (no Item_No)     : {stats['rows_skipped_no_art']:,}")
    if n > 0:
        year_counts: Counter = Counter()
        for r in rows:
            pd = r["posting_date"]
            yr = int(pd[:4]) if isinstance(pd, str) else pd.year
            year_counts[yr] += 1
        print("   Date span by year:")
        for yr in sorted(year_counts):
            print(f"     {yr}  →  {year_counts[yr]:,} rows")

    print()
    print("2. DOC_TYPE HISTOGRAM")
    doc_raw_counter: Counter = Counter()
    for r in rows:
        doc_raw_counter[r["doc_type_raw"] or ""] += 1
    for raw_label, cnt in sorted(doc_raw_counter.items(), key=lambda x: -x[1]):
        mapped = DOC_TYPE_MAP.get(raw_label, "⚠ UNMAPPED")
        flag = "  ← ⚠ UNMAPPED — ADD TO DOC_TYPE_MAP BEFORE --apply" if raw_label not in DOC_TYPE_MAP else ""
        print(f"   {raw_label!r:<35}  → {mapped:<18}  {cnt:>6,}{flag}")
    if stats.get("unmapped_doc_types"):
        print()
        print("   ⚠ UNMAPPED DOC TYPE LABELS (will abort --apply until mapped):")
        for label, cnt in stats["unmapped_doc_types"].items():
            print(f"     {label!r}  → {cnt} rows")
    else:
        print("   ✓ Zero unmapped doc_types.")

    print()
    print("3. CUSTOMER MATCH RATE")
    cust_null = sum(1 for r in rows if r["customer_id_fk"] is None)
    distinct_source_nos: set[str] = {
        r["bc_source_no"] for r in rows if r["bc_source_no"]
    }
    unresolved_source_nos: set[str] = {
        r["bc_source_no"] for r in rows
        if r["customer_id_fk"] is None and r["bc_source_no"]
    }
    resolved_cnt = len(distinct_source_nos) - len(unresolved_source_nos)
    print(f"   Distinct bc_source_no    : {len(distinct_source_nos):,}")
    print(f"   Resolved to customer_id  : {resolved_cnt:,} / {len(distinct_source_nos):,}")
    print(f"   Rows with NULL customer  : {cust_null:,}")
    if unresolved_source_nos:
        print(f"   Unresolved source_nos    : {sorted(unresolved_source_nos)[:20]}")

    print()
    print("4. SKU MATCH RATE")
    sku_null = sum(1 for r in rows if r["sku_id_fk"] is None)
    distinct_skus: set[str] = {r["sku_code_raw"] for r in rows}
    unresolved_skus: set[str] = {r["sku_code_raw"] for r in rows if r["sku_id_fk"] is None}
    resolved_skus_cnt = len(distinct_skus) - len(unresolved_skus)
    print(f"   Distinct SKU codes       : {len(distinct_skus):,}")
    print(f"   Resolved to sku_id_fk    : {resolved_skus_cnt:,} / {len(distinct_skus):,}")
    print(f"   Rows with NULL sku_id_fk : {sku_null:,}")
    if unresolved_skus:
        print(f"   Unresolved SKU codes     : {sorted(unresolved_skus)[:20]}")

    print()
    print("5. SANITY TOTALS")
    total_qty   = sum(r["qty_signed"] for r in rows)
    total_sales = sum(
        r["sales_amount_chf"] for r in rows if r["sales_amount_chf"] is not None
    )
    total_hl = sum(r["hl_resolved"] for r in rows if r["hl_resolved"] is not None)
    print(f"   SUM(qty_signed)         : {float(total_qty):>15,.4f}")
    print(f"   SUM(sales_amount_chf)   : {float(total_sales):>15,.2f} CHF")
    print(f"   SUM(hl_resolved)        : {float(total_hl):>15,.4f} HL")

    print()
    print("=" * 70)


# ── SKU review CSV writer ──────────────────────────────────────────────────────

_REPO_ROOT = Path("/var/www/maltytask")


def write_sku_review_csv(rows: list[dict[str, Any]], out_path: Path | None = None) -> Path:
    """Write bc-ledger-sku-review.csv with one row per unresolved SKU.
    Defaults to <repo_root>/data/bc-ledger-sku-review.csv; pass out_path to override."""
    if out_path is None:
        out_dir = _REPO_ROOT / "data"
        out_dir.mkdir(exist_ok=True)
        out_path = out_dir / "bc-ledger-sku-review.csv"

    unresolved: dict[str, dict[str, Any]] = {}
    for r in rows:
        if r["sku_id_fk"] is not None:
            continue
        code = r["sku_code_raw"]
        d_str = r["posting_date"]
        qty = r["qty_signed"]
        if code not in unresolved:
            unresolved[code] = {
                "row_count":         0,
                "total_qty_signed":  Decimal("0"),
                "first_posting_date": d_str,
                "last_posting_date":  d_str,
            }
        b = unresolved[code]
        b["row_count"] += 1
        b["total_qty_signed"] += qty
        if d_str < b["first_posting_date"]:
            b["first_posting_date"] = d_str
        if d_str > b["last_posting_date"]:
            b["last_posting_date"] = d_str

    sorted_codes = sorted(unresolved.items(), key=lambda x: -x[1]["row_count"])
    fieldnames = [
        "sku_code", "row_count", "total_qty_signed",
        "first_posting_date", "last_posting_date", "decision",
    ]
    with out_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for code, b in sorted_codes:
            writer.writerow({
                "sku_code":           code,
                "row_count":          b["row_count"],
                "total_qty_signed":   str(b["total_qty_signed"]),
                "first_posting_date": str(b["first_posting_date"]),
                "last_posting_date":  str(b["last_posting_date"]),
                "decision":           "",
            })
    return out_path


# ── Overlap verifier (GATE A) ─────────────────────────────────────────────────

_COMPARE_FIELDS = ("doc_type", "qty_signed", "sales_amount_chf",
                   "customer_id_fk", "sku_id_fk", "hl_resolved")


def run_verify_overlap(
    conn: pymysql.connections.Connection,
    bc: dict[str, str],
    from_seq: int,
    to_seq: int,
    customers: dict[str, int],
    active_skus: dict[str, tuple[int, float | None]],
    alias_map: dict[str, tuple[int, float | None]],
) -> bool:
    """
    Pull Entry_No in [from_seq, to_seq] from the BC API, map them with
    parse_api_row, then diff each field in _COMPARE_FIELDS against the
    live DB rows for those Entry_Nos.

    PASS/FAIL semantics:
      PASS — all rows present in BOTH API and DB are field-identical (within
             DECIMAL(14,4) storage precision — the DB column truncates to 4dp,
             but the API returns up to 5dp; we round both to 4dp before comparing).
      FAIL — any row present in both has a field-level mismatch that is NOT
             explained by decimal truncation.  This means the English API field
             mapping disagrees with the historical French xlsx mapping.

    "Missing from DB" entries (present in API but absent from DB) are INFORMATIONAL
    only — they are BC entries posted after the historical xlsx snapshot was taken.
    They are exactly what the incremental connector will load on its first --apply
    run.  They do NOT indicate a mapping problem and do NOT cause a FAIL.

    Read-only: never writes.
    """
    print(f"\nGATE A — verify-overlap [{from_seq:,} … {to_seq:,}]", flush=True)
    print("Fetching range from BC API …", flush=True)
    api_rows_raw = fetch_odata_range(bc, from_seq, to_seq)
    print(f"API returned {len(api_rows_raw)} rows.", flush=True)

    if not api_rows_raw:
        print("WARN: API returned 0 rows for the range — nothing to compare.")
        return True

    # Map to row dicts
    unmapped: Counter = Counter()
    api_mapped: dict[int, dict[str, Any]] = {}
    for r in api_rows_raw:
        parsed = parse_api_row(r, customers, active_skus, alias_map, unmapped)
        if parsed is None:
            continue
        seq = parsed["bc_line_seq"]
        if seq is not None:
            api_mapped[seq] = parsed

    if unmapped:
        print(f"  WARN: unmapped doc_types in range: {dict(unmapped)}")

    # Fetch DB rows for the same Entry_Nos
    seqs = list(api_mapped.keys())
    if not seqs:
        print("WARN: no parseable rows from API in range.")
        return True

    # Chunk IN clause to avoid too-large queries
    db_rows: dict[int, dict[str, Any]] = {}
    chunk_size = 1000
    with conn.cursor() as c:
        for start in range(0, len(seqs), chunk_size):
            chunk = seqs[start: start + chunk_size]
            placeholders = ",".join(["%s"] * len(chunk))
            c.execute(
                f"SELECT bc_line_seq, doc_type, qty_signed, sales_amount_chf, "
                f"       customer_id_fk, sku_id_fk, hl_resolved "
                f"FROM inv_sales_ledger "
                f"WHERE bc_line_seq IN ({placeholders})",
                chunk,
            )
            for row in c.fetchall():
                db_rows[int(row["bc_line_seq"])] = row

    both = set(api_mapped) & set(db_rows)
    missing_in_db = set(api_mapped) - set(db_rows)
    missing_in_api = set(db_rows) - set(api_mapped)

    print(f"  API mapped rows   : {len(api_mapped):,}")
    print(f"  DB rows found     : {len(db_rows):,}")
    print(f"  Common (compared) : {len(both):,}")
    print(f"  In API not in DB  : {len(missing_in_db):,}  "
          f"(informational — new BC entries since historical xlsx snapshot)")
    if missing_in_api:
        print(f"  In DB not in API  : {len(missing_in_api):,}  "
              f"(unexpected — possibly retired/deleted in BC)")
        for seq in sorted(missing_in_api)[:5]:
            print(f"    Entry_No {seq}: in DB but NOT in API")

    # Field-by-field comparison for rows present in BOTH.
    # Decimal comparison: round both sides to 4 decimal places (DECIMAL(14,4)
    # storage precision — the API may return 5+ significant decimals which MySQL
    # truncates on insert; this is expected behaviour, not a mapping error).
    mapping_diffs: list[str] = []
    for seq in sorted(both):
        api_r = api_mapped[seq]
        db_r  = db_rows[seq]
        for field in _COMPARE_FIELDS:
            api_val = api_r.get(field)
            db_val  = db_r.get(field)
            if isinstance(api_val, Decimal) and db_val is not None:
                try:
                    # Round both to 4dp (DECIMAL(14,4) storage)
                    api_4dp = round(api_val, 4)
                    db_4dp  = round(Decimal(str(db_val)), 4)
                    if api_4dp != db_4dp:
                        mapping_diffs.append(
                            f"  Entry_No {seq} field '{field}': "
                            f"API={api_val!r} (rounded={api_4dp}) vs DB={db_val!r}"
                        )
                    continue
                except InvalidOperation:
                    pass
            if api_val != db_val:
                mapping_diffs.append(
                    f"  Entry_No {seq} field '{field}': "
                    f"API={api_val!r} vs DB={db_val!r}"
                )

    print()
    if not mapping_diffs:
        print(
            f"GATE A RESULT: ✓ PASS — zero field-level mapping diffs across "
            f"{len(both):,} compared rows.\n"
            f"  ({len(missing_in_db):,} entries in API not yet in DB — "
            f"expected; the incremental connector will load them.)"
        )
        return True
    else:
        print(
            f"GATE A RESULT: ✗ FAIL — {len(mapping_diffs):,} field-level "
            f"mapping mismatches detected (DECIMAL precision already accounted for):"
        )
        for d in mapping_diffs[:50]:
            print(d)
        if len(mapping_diffs) > 50:
            print(f"  … and {len(mapping_diffs) - 50} more diffs (truncated).")
        print()
        print("STOP: do not proceed to --apply or cron-enable.")
        print("A divergence means the English API field mapping differs from")
        print("the historical French xlsx mapping. Escalate to operator.")
        return False


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    ap = argparse.ArgumentParser(
        description="BC item-ledger → inv_sales_ledger incremental connector."
    )
    ap.add_argument(
        "--source", choices=["api", "xlsx"], default="api",
        help="Data source: 'api' (BC OData, default) or 'xlsx' (one-off reload).",
    )
    ap.add_argument(
        "--file", metavar="PATH",
        help="xlsx path — required when --source xlsx.",
    )
    ap.add_argument(
        "--dry-run", action="store_true", default=True,
        help="Parse + report only, no writes (default).",
    )
    ap.add_argument(
        "--apply", action="store_true", default=False,
        help="Write rows to inv_sales_ledger (overrides --dry-run).",
    )
    ap.add_argument(
        "--limit", type=int, metavar="N",
        help="Process first N rows only (testing).",
    )
    ap.add_argument(
        "--verify-overlap", nargs=2, type=int, metavar=("FROM", "TO"),
        help=(
            "Read-only GATE A check: pull Entry_No in [FROM, TO] from the BC API, "
            "diff field-by-field against DB rows for those Entry_Nos. "
            "Example: --verify-overlap 60000 61500"
        ),
    )
    args = ap.parse_args()

    dry_run = not args.apply

    # ── Connect + load reference data ─────────────────────────────────────────
    cfg  = load_config()
    conn = pymysql.connect(
        host=cfg.db_host,
        port=cfg.db_port,
        user=cfg.db_user,
        password=cfg.db_password,
        database=cfg.db_name,
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor,
        autocommit=False,
        connect_timeout=10,
        read_timeout=120,
        write_timeout=120,
    )
    print("Connected to maltytask MySQL.", flush=True)

    try:
        print("Loading reference data …", end=" ", flush=True)
        customers = load_customers(conn)
        active_skus, alias_map = load_skus(conn)
        print(
            f"done — {len(customers)} customers, "
            f"{len(active_skus)} active SKUs, "
            f"{len(alias_map)} SKU aliases."
        )

        # ── --verify-overlap (GATE A, read-only) ──────────────────────────────
        if args.verify_overlap:
            from_seq, to_seq = args.verify_overlap
            bc = _load_bc_env()
            passed = run_verify_overlap(
                conn, bc, from_seq, to_seq,
                customers, active_skus, alias_map,
            )
            conn.close()
            sys.exit(0 if passed else 1)

        # ── --source xlsx (fallback reload) ───────────────────────────────────
        if args.source == "xlsx":
            if not args.file:
                print("ERROR: --source xlsx requires --file PATH", file=sys.stderr)
                conn.close()
                sys.exit(1)
            xlsx_path = Path(args.file)
            if not xlsx_path.exists():
                print(f"ERROR: file not found: {xlsx_path}", file=sys.stderr)
                conn.close()
                sys.exit(1)
            print(f"Parsing {xlsx_path.name} …", flush=True)
            rows, stats = parse_xlsx(xlsx_path, customers, active_skus, alias_map, args.limit)
            print(f"Parsed {len(rows):,} rows.")

        # ── --source api (default incremental path) ───────────────────────────
        # Incremental strategy: ANTI-JOIN on Entry_No, bounded to the tracked
        # period.  BC does NOT assign Entry_No in strict chronological order, so
        # a `Entry_No gt MAX(bc_line_seq)` watermark permanently misses "holes
        # below the high-water mark" (entries posted after the snapshot but
        # carrying an Entry_No below the current max).  The anti-join finds them:
        #   missing = scoped_api_entry_nos − db_bc_line_seqs
        # then fetches FULL rows only for `missing` and upserts.
        else:
            cutoff = get_scope_cutoff(conn)
            if cutoff is None:
                print(
                    "inv_sales_ledger is empty — no scope cutoff. "
                    "Seed with --source xlsx first.",
                    file=sys.stderr,
                )
                conn.close()
                sys.exit(1)
            print(f"Scope cutoff: MIN(posting_date) = {cutoff}", flush=True)

            bc = _load_bc_env()

            # 1. Cheap projection of all in-scope Sale Entry_Nos.
            scoped_api_seqs = fetch_scoped_entry_nos(bc, cutoff)
            # 2. DB key set.
            db_seqs = load_db_entry_nos(conn)
            # 3. Anti-join — entries present in BC's in-scope set but not in DB.
            missing = scoped_api_seqs - db_seqs

            print(
                f"  [anti-join] in-scope API Entry_Nos: {len(scoped_api_seqs):,} | "
                f"DB Entry_Nos: {len(db_seqs):,} | missing (to insert): {len(missing):,}",
                flush=True,
            )

            if not missing:
                print("No missing entries — DB is in sync with BC. Nothing to do.")
                conn.close()
                return

            # --limit caps the missing set (testing). Take the lowest Entry_Nos.
            missing_list = sorted(missing)
            if args.limit is not None:
                missing_list = missing_list[: args.limit]
                print(f"  [--limit] capped missing set to {len(missing_list):,} Entry_Nos.",
                      flush=True)

            # 4. Fetch FULL rows for the missing set (OR-chained batches).
            api_raw = fetch_odata_by_entry_nos(bc, missing_list)

            unmapped: Counter = Counter()
            rows = []
            for r in api_raw:
                parsed = parse_api_row(r, customers, active_skus, alias_map, unmapped)
                if parsed is not None:
                    rows.append(parsed)
            skipped = len(api_raw) - len(rows)
            stats = {
                "source":              API_SOURCE_FILE,
                "rows_data":           len(rows),
                "rows_skipped_no_art": skipped,
                "unmapped_doc_types":  dict(unmapped),
                "scope_cutoff":        cutoff,
                "missing_count":       len(missing_list),
            }
            print(
                f"Mapped {len(rows):,} rows "
                f"({skipped} skipped — no Item_No).",
                flush=True,
            )

        # ── Dry-run ───────────────────────────────────────────────────────────
        if dry_run:
            build_dry_run_report(rows, stats)
            if rows and any(r["sku_id_fk"] is None for r in rows):
                import tempfile
                try:
                    csv_path = write_sku_review_csv(rows)
                except PermissionError:
                    # /var/www/maltytask/data/ may not be writable by the running user
                    # (owned by ubuntu, not maltytask).  Fall back to /tmp.
                    fallback = Path(tempfile.gettempdir()) / "bc-ledger-sku-review.csv"
                    csv_path = write_sku_review_csv(rows, out_path=fallback)
                print(f"SKU review CSV → {csv_path}")
            if stats.get("unmapped_doc_types"):
                print(
                    "\n⚠ UNMAPPED doc_type labels — add to DOC_TYPE_MAP before --apply.",
                    file=sys.stderr,
                )
            conn.close()
            print("\nDry-run complete. Run with --apply to write to DB.")
            return

        # ── Apply: abort on unmapped doc_types ────────────────────────────────
        if stats.get("unmapped_doc_types"):
            print(
                "ERROR: unmapped doc_type labels — cannot --apply.\n"
                "Add them to DOC_TYPE_MAP first:",
                file=sys.stderr,
            )
            for label, cnt in stats["unmapped_doc_types"].items():
                print(f"  {label!r}  ({cnt} rows)", file=sys.stderr)
            conn.close()
            sys.exit(1)

        # ── Apply: write ──────────────────────────────────────────────────────
        print(f"Writing {len(rows):,} rows to inv_sales_ledger …", flush=True)
        rc = insert_rows(conn, rows)
        print(
            f"Done — DB rowcount signal: {rc} "
            f"(1=new, 2=updated, 0=unchanged per row; total rows: {len(rows)})"
        )
        conn.close()
        print("\nApply complete.")

    except Exception:
        conn.close()
        raise


if __name__ == "__main__":
    main()
