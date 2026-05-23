"""
ingest_bd_brewing_v2.py — upload normalized BrewingData into bd_brewing_*_v2 tables.

Source: /var/www/maltytask/data/RawDB-normalized.xlsx
Sheets:
  BrewingData_Brewday       → bd_brewing_brewday_v2     (1 row per batch)
  BrewingData_FirstWort     ┐
  BrewingData_Pfannevoll    ├→ bd_brewing_gravity_v2    (folded via event_type)
  BrewingData_Kochwurze     │
  BrewingData_Cooling       ┘
  BrewingData_Timings       → bd_brewing_timings_v2     (1 row per brew)
  BrewingData_Ingredients   → bd_brewing_ingredients_v2 (header, 1 per batch)
                            + bd_brewing_ingredients_parsed_v2 (child, 1 per parsed line)

OPERATOR SETUP: Before running, SCP the xlsx from the local maltytask machine:
  scp /home/kluk/projects/maltytask/data/RawDB-normalized.xlsx \\
      maltyweb:/var/www/maltytask/data/RawDB-normalized.xlsx

Usage:
  python3 ingest_bd_brewing_v2.py                        # dry-run all tables
  python3 ingest_bd_brewing_v2.py --apply                # live write all tables
  python3 ingest_bd_brewing_v2.py --table brewday        # dry-run brewday only
  python3 ingest_bd_brewing_v2.py --table gravity        # dry-run gravity only
  python3 ingest_bd_brewing_v2.py --table timings        # dry-run timings only
  python3 ingest_bd_brewing_v2.py --table ingredients    # dry-run ingredients only
  python3 ingest_bd_brewing_v2.py --table brewday --apply --limit 10

Tables:  brewday | gravity | timings | ingredients  (default: all)
"""
from __future__ import annotations

import argparse
import hashlib
import json
import sys
from collections import Counter
from datetime import datetime, date, time
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

_SCRIPT_DIR = Path(__file__).parent
if str(_SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPT_DIR))

from lib_config import load as load_config
from lib_db import connect

# ── Constants ──────────────────────────────────────────────────────────────────

XLSX_PATH = Path("/var/www/maltytask/data/RawDB-normalized.xlsx")

# ── Type coercions ─────────────────────────────────────────────────────────────

def _s(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _i(v: Any) -> int | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    if not s:
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def _f(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return float(v)
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _dt(v: Any) -> str | None:
    """Coerce to MySQL DATETIME(6) string or None."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S.%f")
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day).strftime("%Y-%m-%d %H:%M:%S.000000")
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d %H:%M:%S.%f")
        except ValueError:
            pass
    return None


def _date(v: Any) -> str | None:
    """Coerce to MySQL DATE string or None."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def _combine_dt(base_dt: Any, t: Any) -> str | None:
    """
    Combine a base datetime/date with a time object to produce DATETIME string.
    brew_start/brew_end in the xlsx are datetime.time objects; the base date comes
    from the Timestamp column (col 0) of the same row.
    """
    if t is None:
        return None
    base_date: date | None = None
    if isinstance(base_dt, datetime):
        base_date = base_dt.date()
    elif isinstance(base_dt, date):
        base_date = base_dt
    if base_date is None:
        return None
    if isinstance(t, time):
        return datetime(base_date.year, base_date.month, base_date.day,
                        t.hour, t.minute, t.second).strftime("%Y-%m-%d %H:%M:%S.000000")
    # fallback: already a datetime
    return _dt(t)


def _row_hash(d: dict) -> str:
    return hashlib.sha256(
        json.dumps(d, sort_keys=True, default=str).encode("utf-8")
    ).hexdigest()


def _is_empty_row(row: tuple) -> bool:
    return not any(v is not None and str(v).strip() != "" for v in row)


# ── Lookup helpers ─────────────────────────────────────────────────────────────

def load_recipe_map(conn) -> dict[int, int]:
    """xlsx beer_recipe_id (same as ref_recipes.id) — validate they exist."""
    with conn.cursor() as cur:
        cur.execute("SELECT id FROM ref_recipes")
        return {r["id"]: r["id"] for r in cur.fetchall()}


def load_mi_string_map(conn) -> dict[str, int]:
    """mi_id (string) → ref_mi.id (int)."""
    with conn.cursor() as cur:
        cur.execute("SELECT id, mi_id FROM ref_mi WHERE mi_id IS NOT NULL")
        return {r["mi_id"]: int(r["id"]) for r in cur.fetchall()}


def load_valid_cct(conn) -> set[int]:
    with conn.cursor() as cur:
        cur.execute("SELECT number FROM ref_cct")
        return {r["number"] for r in cur.fetchall()}


def load_valid_yt(conn) -> set[int]:
    with conn.cursor() as cur:
        cur.execute("SELECT number FROM ref_yt")
        return {r["number"] for r in cur.fetchall()}


# ── UPSERT helper ──────────────────────────────────────────────────────────────

def upsert_rows(conn, table: str, rows: list[dict], natural_keys: frozenset[str],
                dry_run: bool, label: str) -> Counter:
    """
    INSERT ... ON DUPLICATE KEY UPDATE for a list of dicts.
    Natural-key columns are excluded from the UPDATE clause.
    row_hash and imported_at are also excluded from UPDATE.
    """
    if not rows:
        print(f"  [{label}] 0 rows — nothing to upsert")
        return Counter(inserted=0, updated=0, skipped=0)

    cols = list(rows[0].keys())
    placeholders = ", ".join(["%s"] * len(cols))
    col_names = ", ".join(f"`{c}`" for c in cols)

    update_cols = [c for c in cols if c not in natural_keys
                   and c not in ("row_hash", "imported_at", "id")]
    if not update_cols:
        update_clause = "row_hash=row_hash"  # no-op — handles edge case
    else:
        update_clause = ", ".join(f"`{c}`=VALUES(`{c}`)" for c in update_cols)

    sql = (
        f"INSERT INTO `{table}` ({col_names}) VALUES ({placeholders})\n"
        f"ON DUPLICATE KEY UPDATE {update_clause}"
    )

    counts = Counter(inserted=0, updated=0)
    if dry_run:
        print(f"  [{label}] DRY-RUN: would upsert {len(rows)} rows into {table}")
        # Show a sample
        for r in rows[:3]:
            print(f"    sample: {json.dumps({k: str(v)[:40] for k, v in r.items() if v is not None}, indent=None)}")
        return Counter(inserted=len(rows), updated=0)

    batch_size = 500
    with conn.cursor() as cur:
        for start in range(0, len(rows), batch_size):
            batch = rows[start:start + batch_size]
            values = [[r[c] for c in cols] for r in batch]
            cur.executemany(sql, values)
            counts["affected"] += cur.rowcount
    conn.commit()
    # MySQL rowcount: 1=inserted, 2=updated, 0=no change (same data)
    print(f"  [{label}] upserted {len(rows)} rows (MySQL affected rows: {counts['affected']})")
    return counts


# ═══════════════════════════════════════════════════════════════════════════════
# TABLE: bd_brewing_brewday_v2
# ═══════════════════════════════════════════════════════════════════════════════

BREWDAY_NK = frozenset({"beer", "batch"})


def ingest_brewday(conn, dry_run: bool, limit: int | None, recipe_map: dict,
                   valid_cct: set, valid_yt: set) -> Counter:
    """Upload BrewingData_Brewday → bd_brewing_brewday_v2."""
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb["BrewingData_Brewday"]

    rows_out: list[dict] = []
    stats = Counter()

    for xlsx_row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if _is_empty_row(row):
            continue
        if limit is not None and len(rows_out) >= limit:
            break

        # Col indices: Timestamp(0) Email(1) Beer(2) Batch(3) beer_recipe_id(4)
        #              CCT(5) CCT_CIP(6) Date_of_CCT_CIP(7) Yeast(8) Yeast_Gen(9)
        #              New_Yeast(10) Pitched_From(11) YT_Number(12) Date_of_YT_CIP(13)
        r = list(row) + [None] * max(0, 14 - len(row))

        beer  = _s(r[2])
        # NULL batch → empty string sentinel so NOT NULL constraint is satisfied.
        # Historic contract brew records occasionally lack a batch number.
        batch = str(_i(r[3])) if r[3] is not None else ""
        if not beer:
            stats["skipped_no_beer"] += 1
            continue

        recipe_id_raw = _i(r[4])
        audit_flags: list[str] = []
        if batch == "":
            audit_flags.append("batch_null")

        recipe_id_fk: int | None = None
        if recipe_id_raw is not None and recipe_id_raw in recipe_map:
            recipe_id_fk = recipe_id_raw
        elif recipe_id_raw is not None:
            audit_flags.append(f"recipe_id_not_found:{recipe_id_raw}")

        cct = _i(r[5])
        if cct is not None and cct not in valid_cct:
            audit_flags.append(f"cct_not_found:{cct}")
            cct = None

        yt_number = _i(r[12])
        if yt_number is not None and yt_number not in valid_yt:
            audit_flags.append(f"yt_not_found:{yt_number}")
            yt_number = None

        canonical = {
            "beer": beer,
            "batch": batch,
            "recipe_id_fk": recipe_id_fk,
            "submitted_at": _dt(r[0]),
            "email": _s(r[1]),
            "event_date": _date(r[0]),
            "cct": cct,
            "cct_cip": _s(r[6]),
            "cct_cip_date": _s(r[7]),
            "yeast": _s(r[8]),
            "yeast_gen": _s(r[9]),
            "new_yeast": _s(r[10]),
            "pitched_from": _s(r[11]),
            "yt_number": yt_number,
            "yt_cip_date": _s(r[13]),
            "start_ferm": None,  # not in Brewday sheet
        }

        rows_out.append({
            **canonical,
            "row_hash": _row_hash(canonical),
            "is_tombstoned": 0,
            "audit_flags": ",".join(audit_flags) if audit_flags else None,
        })
        stats["rows_parsed"] += 1

    wb.close()
    print(f"  [brewday] parsed {stats['rows_parsed']} rows "
          f"(skipped no-beer: {stats['skipped_no_beer']})")

    upsert_rows(conn, "bd_brewing_brewday_v2", rows_out, BREWDAY_NK, dry_run, "brewday")
    return stats


# ═══════════════════════════════════════════════════════════════════════════════
# TABLE: bd_brewing_gravity_v2
# ═══════════════════════════════════════════════════════════════════════════════

GRAVITY_NK = frozenset({"beer", "batch", "brew", "event_type"})

_GRAVITY_SHEETS = [
    ("BrewingData_FirstWort",  "FirstWort"),
    ("BrewingData_Pfannevoll", "Pfannevoll"),
    ("BrewingData_Kochwurze",  "Kochwurze"),
    ("BrewingData_Cooling",    "Cooling"),
]


def _parse_gravity_row(row: list, event_type: str, recipe_map: dict) -> dict | None:
    """
    Parse one xlsx gravity row.
    All sheets share: Timestamp(0) Email(1) Beer(2) Batch(3) beer_recipe_id(4) Brew(5)
    Then event-specific cols from index 6 onward.
    Returns None if beer is blank.
    """
    r = list(row) + [None] * max(0, 10 - len(row))

    beer  = _s(r[2])
    if not beer:
        return None
    batch = str(_i(r[3])) if r[3] is not None else ""
    brew  = str(_i(r[5])) if r[5] is not None else ""

    recipe_id_raw = _i(r[4])
    audit_flags: list[str] = []
    if batch == "":
        audit_flags.append("batch_null")
    if brew == "":
        audit_flags.append("brew_null")

    recipe_id_fk: int | None = None
    if recipe_id_raw is not None and recipe_id_raw in recipe_map:
        recipe_id_fk = recipe_id_raw
    elif recipe_id_raw is not None:
        audit_flags.append(f"recipe_id_not_found:{recipe_id_raw}")

    # Event-specific fields (all default None)
    fw_gravity = fw_ph = pfv_gravity = kw_gravity = None
    final_ph = final_gravity = final_volume = batch_dilution = None

    if event_type == "FirstWort":
        fw_gravity = _f(r[6])
        fw_ph      = _f(r[7])
    elif event_type == "Pfannevoll":
        pfv_gravity = _f(r[6])
    elif event_type == "Kochwurze":
        kw_gravity  = _f(r[6])
    elif event_type == "Cooling":
        final_ph       = _f(r[6])
        final_gravity  = _f(r[7])
        final_volume   = _f(r[8])
        batch_dilution = _f(r[9]) if len(r) > 9 else None

    canonical = {
        "beer": beer,
        "batch": batch,
        "brew": brew,
        "event_type": event_type,
        "recipe_id_fk": recipe_id_fk,
        "submitted_at": _dt(r[0]),
        "email": _s(r[1]),
        "firstwort_gravity": fw_gravity,
        "firstwort_ph": fw_ph,
        "pfannevoll_gravity": pfv_gravity,
        "kochwurze_gravity": kw_gravity,
        "final_ph": final_ph,
        "final_gravity": final_gravity,
        "final_volume": final_volume,
        "batch_dilution": batch_dilution,
    }

    return {
        **canonical,
        "row_hash": _row_hash(canonical),
        "is_tombstoned": 0,
        "audit_flags": ",".join(audit_flags) if audit_flags else None,
    }


def ingest_gravity(conn, dry_run: bool, limit: int | None, recipe_map: dict) -> Counter:
    """Upload 4 gravity sheets → bd_brewing_gravity_v2 (folded by event_type)."""
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

    rows_out: list[dict] = []
    stats = Counter()

    for sheet_name, event_type in _GRAVITY_SHEETS:
        ws = wb[sheet_name]
        sheet_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if _is_empty_row(row):
                continue
            if limit is not None and len(rows_out) >= limit:
                break
            parsed = _parse_gravity_row(list(row), event_type, recipe_map)
            if parsed is None:
                stats["skipped_no_beer"] += 1
                continue
            rows_out.append(parsed)
            stats["rows_parsed"] += 1
            sheet_count += 1
        print(f"  [gravity] {sheet_name}: {sheet_count} rows")

    wb.close()
    print(f"  [gravity] total parsed: {stats['rows_parsed']}")
    upsert_rows(conn, "bd_brewing_gravity_v2", rows_out, GRAVITY_NK, dry_run, "gravity")
    return stats


# ═══════════════════════════════════════════════════════════════════════════════
# TABLE: bd_brewing_timings_v2
# ═══════════════════════════════════════════════════════════════════════════════

TIMINGS_NK = frozenset({"beer", "batch", "brew"})


def ingest_timings(conn, dry_run: bool, limit: int | None, recipe_map: dict) -> Counter:
    """Upload BrewingData_Timings → bd_brewing_timings_v2."""
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb["BrewingData_Timings"]

    rows_out: list[dict] = []
    stats = Counter()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if _is_empty_row(row):
            continue
        if limit is not None and len(rows_out) >= limit:
            break

        # Timestamp(0) Email(1) Beer(2) Batch(3) beer_recipe_id(4) Brew(5)
        # Brew_Start(6) Brew_End(7)
        r = list(row) + [None] * max(0, 8 - len(row))

        beer  = _s(r[2])
        if not beer:
            stats["skipped_no_beer"] += 1
            continue
        batch = str(_i(r[3])) if r[3] is not None else ""
        brew  = str(_i(r[5])) if r[5] is not None else ""

        recipe_id_raw = _i(r[4])
        audit_flags: list[str] = []
        if batch == "":
            audit_flags.append("batch_null")
        if brew == "":
            audit_flags.append("brew_null")

        recipe_id_fk: int | None = None
        if recipe_id_raw is not None and recipe_id_raw in recipe_map:
            recipe_id_fk = recipe_id_raw
        elif recipe_id_raw is not None:
            audit_flags.append(f"recipe_id_not_found:{recipe_id_raw}")

        # brew_start/brew_end are time objects — combine with Timestamp date
        brew_start = _combine_dt(r[0], r[6])
        brew_end   = _combine_dt(r[0], r[7])
        event_date = _date(r[0])

        canonical = {
            "beer": beer,
            "batch": batch,
            "brew": brew,
            "recipe_id_fk": recipe_id_fk,
            "submitted_at": _dt(r[0]),
            "email": _s(r[1]),
            "brew_start": brew_start,
            "brew_end": brew_end,
            "event_date": event_date,
            "start_ferm": None,
        }

        rows_out.append({
            **canonical,
            "row_hash": _row_hash(canonical),
            "is_tombstoned": 0,
            "audit_flags": ",".join(audit_flags) if audit_flags else None,
        })
        stats["rows_parsed"] += 1

    wb.close()
    print(f"  [timings] parsed {stats['rows_parsed']} rows")
    upsert_rows(conn, "bd_brewing_timings_v2", rows_out, TIMINGS_NK, dry_run, "timings")
    return stats


# ═══════════════════════════════════════════════════════════════════════════════
# TABLE: bd_brewing_ingredients_v2 + bd_brewing_ingredients_parsed_v2
# ═══════════════════════════════════════════════════════════════════════════════

INGREDIENTS_NK     = frozenset({"beer", "batch"})
INGREDIENTS_PSD_NK = frozenset({"header_id", "category", "line_idx"})


def ingest_ingredients(conn, dry_run: bool, limit: int | None, recipe_map: dict,
                       mi_string_map: dict) -> Counter:
    """
    Upload BrewingData_Ingredients.
    Each xlsx row is a parsed ingredient line.  We first group by (beer, batch) to
    build header rows, then insert headers, then insert child (parsed) rows using
    the header_id returned by the INSERT.

    Col map:
      timestamp(0) email(1) event_date(2) beer_raw(3) batch(4) beer_recipe_id(5)
      category(6) line_idx(7) mi_id_resolved(8) raw_name(9) qty(10) unit(11)
      lot(12) confidence(13) parse_note(14) source_row(15)
    """
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb["BrewingData_Ingredients"]

    # Group xlsx lines by (beer, batch) natural key
    from collections import defaultdict
    groups: dict[tuple[str | None, str | None], list] = defaultdict(list)
    group_meta: dict[tuple, dict] = {}  # per (beer,batch): first-seen provenance

    total_lines = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if _is_empty_row(row):
            continue
        r = list(row) + [None] * max(0, 16 - len(row))

        beer  = _s(r[3]) or _s(r[2])   # beer_raw(3) preferred
        batch_val = _i(r[4])
        batch = str(batch_val) if batch_val is not None else ""
        key = (beer, batch)

        if key not in group_meta:
            group_meta[key] = {
                "submitted_at": _dt(r[0]),
                "email": _s(r[1]),
                "event_date": _date(r[2]) or _date(r[0]),
                "beer_recipe_id": _i(r[5]),
            }
        groups[key].append(r)
        total_lines += 1

    wb.close()
    print(f"  [ingredients] {len(groups)} header groups, {total_lines} parsed lines")

    stats = Counter()
    header_natural_keys = list(groups.keys())
    if limit is not None:
        header_natural_keys = header_natural_keys[:limit]

    # ── Pass 1: upsert headers ─────────────────────────────────────────────────
    header_rows: list[dict] = []
    for (beer, batch) in header_natural_keys:
        if not beer:
            stats["skipped_no_beer"] += 1
            continue
        meta = group_meta[(beer, batch)]
        recipe_id_raw = meta["beer_recipe_id"]
        # batch="" sentinel (see grouping above) — flag it
        audit_flags: list[str] = []
        if batch == "":
            audit_flags.append("batch_null")

        recipe_id_fk: int | None = None
        if recipe_id_raw is not None and recipe_id_raw in recipe_map:
            recipe_id_fk = recipe_id_raw
        elif recipe_id_raw is not None:
            audit_flags.append(f"recipe_id_not_found:{recipe_id_raw}")

        canonical = {
            "beer": beer,
            "batch": batch,
            "recipe_id_fk": recipe_id_fk,
            "submitted_at": meta["submitted_at"],
            "email": meta["email"],
            "event_date": meta["event_date"],
            "raw_blob_text": None,  # not stored in xlsx
            "parsed_at": None,
        }
        header_rows.append({
            **canonical,
            "row_hash": _row_hash(canonical),
            "is_tombstoned": 0,
            "audit_flags": ",".join(audit_flags) if audit_flags else None,
        })
        stats["headers_parsed"] += 1

    if dry_run:
        print(f"  [ingredients] DRY-RUN: would upsert {len(header_rows)} header rows "
              f"and ~{total_lines} parsed lines into bd_brewing_ingredients_v2 "
              f"/ bd_brewing_ingredients_parsed_v2")
        for r in header_rows[:3]:
            print(f"    header sample: {r['beer']!r} batch={r['batch']!r}")
        return stats

    # Live: upsert headers, then look up their IDs for child insert
    upsert_rows(conn, "bd_brewing_ingredients_v2", header_rows, INGREDIENTS_NK,
                dry_run=False, label="ingredients-header")

    # Load header_id map (beer, batch) → id
    with conn.cursor() as cur:
        cur.execute("SELECT id, beer, batch FROM bd_brewing_ingredients_v2")
        header_id_map: dict[tuple, int] = {
            (r["beer"], r["batch"]): r["id"] for r in cur.fetchall()
        }

    # ── Pass 2: upsert parsed lines ────────────────────────────────────────────
    parsed_rows: list[dict] = []
    for (beer, batch) in header_natural_keys:
        header_id = header_id_map.get((beer, batch))
        if header_id is None:
            stats["skipped_no_header_id"] += 1
            continue
        for r in groups[(beer, batch)]:
            mi_str = _s(r[8])  # mi_id_resolved
            mi_id_fk = mi_string_map.get(mi_str) if mi_str else None
            if mi_str and mi_id_fk is None:
                audit_flags = f"mi_unresolved:{mi_str}"
            else:
                audit_flags = None

            cat_raw = _s(r[6])
            # Map to ENUM: malt | hops_kettle | hops_dry
            if cat_raw == "malt":
                category = "malt"
            elif cat_raw in ("hops_kettle", "hops"):
                category = "hops_kettle"
            elif cat_raw == "hops_dry":
                category = "hops_dry"
            else:
                category = "malt"  # safe fallback; audit_flags will note it
                if audit_flags:
                    audit_flags += f",unknown_category:{cat_raw}"
                else:
                    audit_flags = f"unknown_category:{cat_raw}"

            unit_raw = _s(r[11])
            unit = unit_raw if unit_raw in ("kg", "g") else None

            parsed_rows.append({
                "header_id": header_id,
                "line_idx":  _i(r[7]) or 0,
                "category":  category,
                "mi_id_fk":  mi_id_fk,
                "raw_name":  _s(r[9]) or "",
                "qty":       _f(r[10]),
                "unit":      unit,
                "lot":       _s(r[12]),
                "confidence": _s(r[13]),
                "parse_note": _s(r[14]),
                "source_row": _i(r[15]),
                # no row_hash on child table (NK = header_id + line_idx)
            })
            stats["lines_parsed"] += 1

    # Insert parsed rows in batches (no row_hash UNIQUE — NK is (header_id, line_idx))
    if parsed_rows:
        cols = list(parsed_rows[0].keys())
        placeholders = ", ".join(["%s"] * len(cols))
        col_names = ", ".join(f"`{c}`" for c in cols)
        update_cols = [c for c in cols if c not in ("header_id", "category", "line_idx", "id", "imported_at")]
        update_clause = ", ".join(f"`{c}`=VALUES(`{c}`)" for c in update_cols)
        sql = (
            f"INSERT INTO `bd_brewing_ingredients_parsed_v2` ({col_names}) "
            f"VALUES ({placeholders})\n"
            f"ON DUPLICATE KEY UPDATE {update_clause}"
        )
        batch_size = 500
        affected = 0
        with conn.cursor() as cur:
            for start in range(0, len(parsed_rows), batch_size):
                batch = parsed_rows[start:start + batch_size]
                values = [[r[c] for c in cols] for r in batch]
                cur.executemany(sql, values)
                affected += cur.rowcount
        conn.commit()
        print(f"  [ingredients-parsed] upserted {len(parsed_rows)} parsed lines "
              f"(MySQL affected: {affected})")
    else:
        print(f"  [ingredients-parsed] 0 parsed lines to insert")

    return stats


# ═══════════════════════════════════════════════════════════════════════════════
# Verification queries
# ═══════════════════════════════════════════════════════════════════════════════

def run_verification(conn) -> None:
    checks = [
        ("brewday row count",
         "SELECT COUNT(*) AS n FROM bd_brewing_brewday_v2"),
        ("brewday recipe_id_fk NULL",
         "SELECT COUNT(*) AS n FROM bd_brewing_brewday_v2 WHERE recipe_id_fk IS NULL"),
        ("brewday batch NULL",
         "SELECT COUNT(*) AS n FROM bd_brewing_brewday_v2 WHERE batch IS NULL"),
        ("gravity row count",
         "SELECT COUNT(*) AS n FROM bd_brewing_gravity_v2"),
        ("gravity per event_type",
         "SELECT event_type, COUNT(*) AS n FROM bd_brewing_gravity_v2 GROUP BY event_type ORDER BY event_type"),
        ("gravity recipe_id_fk NULL",
         "SELECT COUNT(*) AS n FROM bd_brewing_gravity_v2 WHERE recipe_id_fk IS NULL"),
        ("timings row count",
         "SELECT COUNT(*) AS n FROM bd_brewing_timings_v2"),
        ("timings recipe_id_fk NULL",
         "SELECT COUNT(*) AS n FROM bd_brewing_timings_v2 WHERE recipe_id_fk IS NULL"),
        ("ingredients header count",
         "SELECT COUNT(*) AS n FROM bd_brewing_ingredients_v2"),
        ("ingredients parsed count",
         "SELECT COUNT(*) AS n FROM bd_brewing_ingredients_parsed_v2"),
        ("ingredients mi_id_fk NULL",
         "SELECT COUNT(*) AS n FROM bd_brewing_ingredients_parsed_v2 WHERE mi_id_fk IS NULL"),
        ("ingredients per category",
         "SELECT category, COUNT(*) AS n FROM bd_brewing_ingredients_parsed_v2 GROUP BY category ORDER BY category"),
    ]
    print("\n=== Verification ===")
    with conn.cursor() as cur:
        for label, sql in checks:
            cur.execute(sql)
            rows = cur.fetchall()
            print(f"\n  {label}:")
            for r in rows:
                print(f"    {dict(r)}")


# ═══════════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--apply",  action="store_true",
                        help="Write to DB. Default is dry-run.")
    parser.add_argument("--table", choices=["brewday", "gravity", "timings", "ingredients"],
                        default=None, help="Upload only this table (default: all).")
    parser.add_argument("--limit", type=int, default=None,
                        help="Process only the first N source rows / header groups.")
    parser.add_argument("--verify", action="store_true",
                        help="Run verification queries after upload (live only).")
    args = parser.parse_args()

    dry_run = not args.apply
    tables  = [args.table] if args.table else ["brewday", "gravity", "timings", "ingredients"]

    if not XLSX_PATH.exists():
        print(f"ERROR: xlsx not found at {XLSX_PATH}", file=sys.stderr)
        print("  Run: scp /home/kluk/projects/maltytask/data/RawDB-normalized.xlsx "
              "maltyweb:/var/www/maltytask/data/RawDB-normalized.xlsx", file=sys.stderr)
        sys.exit(1)

    cfg  = load_config()
    conn = connect(cfg)

    print(f"{'DRY-RUN' if dry_run else 'LIVE'} — tables: {', '.join(tables)}")
    print(f"Source: {XLSX_PATH}")
    print()

    recipe_map    = load_recipe_map(conn)
    valid_cct     = load_valid_cct(conn)
    valid_yt      = load_valid_yt(conn)
    mi_string_map = load_mi_string_map(conn)

    print(f"Loaded: {len(recipe_map)} recipes, {len(valid_cct)} CCTs, "
          f"{len(valid_yt)} YTs, {len(mi_string_map)} MI string IDs")
    print()

    if "brewday" in tables:
        print("── Brewday ──────────────────────────────────────────────────────")
        ingest_brewday(conn, dry_run, args.limit, recipe_map, valid_cct, valid_yt)
        print()

    if "gravity" in tables:
        print("── Gravity ──────────────────────────────────────────────────────")
        ingest_gravity(conn, dry_run, args.limit, recipe_map)
        print()

    if "timings" in tables:
        print("── Timings ──────────────────────────────────────────────────────")
        ingest_timings(conn, dry_run, args.limit, recipe_map)
        print()

    if "ingredients" in tables:
        print("── Ingredients ──────────────────────────────────────────────────")
        ingest_ingredients(conn, dry_run, args.limit, recipe_map, mi_string_map)
        print()

    if not dry_run and args.verify:
        run_verification(conn)

    conn.close()
    print("Done.")


if __name__ == "__main__":
    main()
