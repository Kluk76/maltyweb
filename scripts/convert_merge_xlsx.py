#!/usr/bin/env python3
"""
scripts/convert_merge_xlsx.py
Runs LOCALLY (WSL) — converts the operator-filled merge-decisions xlsx into a
decisions JSON consumed by apply_merge_decisions.php on the VPS.

Usage:
  python3 scripts/convert_merge_xlsx.py \
      "/mnt/c/Users/Kouros/Desktop/merge-clients-review.xlsx" \
      /tmp/merge-decisions.json

Output JSON:
  {
    "exported_at": "<iso>",
    "row_count":   N,
    "errors":      ["..."],          # parse-time issues (missing id, unknown sheet, etc.)
    "decisions": [
      {
        "sheet":        "A"|"B"|"C"|"D",
        "ref":          str,
        "id":           int,           # ref_customers.id of the needs_review row
        "name":         str,
        "decision":     "MERGE"|"MERGE_BC"|"VALIDER"|"DÉSACTIVER"|"NON"|"SKIP",
        "suggested_bc": str|null,      # BC n° from the sheet's BC column (A/B only)
        "target_bc":    str|null       # explicit BC n° when operator typed one
      },
      ...
    ]
  }

Decision parsing rules per sheet:
  Sheet A (Fusions proposées) — DÉCISION column:
    OUI (or blank)  → decision=MERGE, target_bc=null   (use suggested_bc from BC col)
    a BC number     → decision=MERGE_BC, target_bc=<that number>
    NON             → decision=NON (skip)

  Sheet B (Incertains) — DÉCISION column:
    OUI             → decision=MERGE, target_bc=null   (use suggested_bc from BC col)
    a BC number     → decision=MERGE_BC, target_bc=<that number>
    VALIDER         → decision=VALIDER
    NON or blank    → decision=NON (skip)

  Sheet C (Privés) — DÉCISION column:
    VALIDER (or blank) → decision=VALIDER
    NON                → decision=NON (skip)

  Sheet D (Puzzles) — DÉCISION column:
    a BC number  → decision=MERGE_BC, target_bc=<that number>
    VALIDER      → decision=VALIDER
    DÉSACTIVER   → decision=DÉSACTIVER
    blank or NON → decision=NON (skip)
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

# ── Sheet metadata ────────────────────────────────────────────────────────────

SHEET_MAP = {
    "A-Fusions proposées": "A",
    "B-Incertains":        "B",
    "C-Privés":            "C",
    "D-Puzzles":           "D",
}

# Column names per sheet (header row 1, case-insensitive match by prefix)
# We locate columns by scanning row 1 for these header strings.
REQUIRED_COLS = {
    "A": ["ref", "id", "nom", "→ fusion vers", "bc", "ville", "score", "2e candidat", "décision"],
    "B": ["ref", "id", "nom", "candidat possible", "bc", "ville", "score", "décision"],
    "C": ["ref", "id", "nom", "décision"],
    "D": ["ref", "id", "nom", "décision"],
}

# Sheets that carry a BC suggestion column
BC_COL_SHEETS = {"A", "B"}

# ── Helpers ───────────────────────────────────────────────────────────────────

_BC_RE = re.compile(r"^\d{4,6}$")


def clean(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def is_bc_number(s: str) -> bool:
    return bool(_BC_RE.match(s))


def find_col(headers: list[str], label: str) -> int | None:
    """Return 0-based index of the first header that starts with label (case-insensitive)."""
    label_lc = label.lower()
    for i, h in enumerate(headers):
        if h.lower().startswith(label_lc):
            return i
    return None


def parse_decision(sheet_key: str, raw: str, suggested_bc: str | None) -> dict:
    """
    Returns dict with keys: decision, target_bc
    decision is one of: MERGE, MERGE_BC, VALIDER, DÉSACTIVER, NON
    """
    v = raw.strip().upper()
    # Strip accents for comparison of VALIDER/DÉSACTIVER
    v_norm = v.replace("É", "E").replace("È", "E").replace("Â", "A").replace("Î", "I")

    if sheet_key == "A":
        if v in ("OUI", "") and suggested_bc:
            return {"decision": "MERGE", "target_bc": None}
        if is_bc_number(v):
            return {"decision": "MERGE_BC", "target_bc": v}
        if v == "NON":
            return {"decision": "NON", "target_bc": None}
        # OUI but no suggested_bc → treat as NON (can't merge without a target)
        if v in ("OUI", ""):
            return {"decision": "NON", "target_bc": None}
        return {"decision": "NON", "target_bc": None}

    if sheet_key == "B":
        if v == "OUI":
            if suggested_bc:
                return {"decision": "MERGE", "target_bc": None}
            return {"decision": "NON", "target_bc": None}
        if is_bc_number(v):
            return {"decision": "MERGE_BC", "target_bc": v}
        if v_norm in ("VALIDER", "VALIDER"):
            return {"decision": "VALIDER", "target_bc": None}
        if v == "NON" or v == "":
            return {"decision": "NON", "target_bc": None}
        return {"decision": "NON", "target_bc": None}

    if sheet_key == "C":
        if v in ("VALIDER", "") or v_norm == "VALIDER":
            return {"decision": "VALIDER", "target_bc": None}
        if v == "NON":
            return {"decision": "NON", "target_bc": None}
        return {"decision": "NON", "target_bc": None}

    if sheet_key == "D":
        if is_bc_number(v):
            return {"decision": "MERGE_BC", "target_bc": v}
        if v_norm in ("VALIDER",):
            return {"decision": "VALIDER", "target_bc": None}
        if v_norm in ("DESACTIVER", "DÉSACTIVER") or v == "DÉSACTIVER":
            return {"decision": "DÉSACTIVER", "target_bc": None}
        if v == "NON" or v == "":
            return {"decision": "NON", "target_bc": None}
        return {"decision": "NON", "target_bc": None}

    return {"decision": "NON", "target_bc": None}


# ── Main ─────────────────────────────────────────────────────────────────────

def parse_sheet(ws, sheet_key: str, parse_errors: list[str]) -> list[dict]:
    rows_out = []
    # Read header row (row 1)
    headers = [clean(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]

    def ci(label: str) -> int | None:
        idx = find_col(headers, label)
        return idx  # 0-based

    col_ref  = ci("ref")
    col_id   = ci("id")
    col_nom  = ci("nom")
    col_dec  = ci("décision")
    col_bc   = ci("bc") if sheet_key in BC_COL_SHEETS else None

    # Validate required columns present
    missing = []
    if col_ref is None: missing.append("ref")
    if col_id  is None: missing.append("id")
    if col_nom is None: missing.append("nom")
    if col_dec is None: missing.append("décision")
    if missing:
        parse_errors.append(f"Sheet {sheet_key}: missing columns: {', '.join(missing)}")
        return rows_out

    for row_idx in range(2, ws.max_row + 1):
        ref_val = clean(ws.cell(row_idx, col_ref + 1).value)
        id_val  = clean(ws.cell(row_idx, col_id  + 1).value)
        nom_val = clean(ws.cell(row_idx, col_nom  + 1).value)
        dec_val = clean(ws.cell(row_idx, col_dec  + 1).value)

        # Skip blank rows
        if not id_val and not nom_val:
            continue

        # Parse id
        try:
            row_id = int(float(id_val)) if id_val else None
        except (ValueError, TypeError):
            parse_errors.append(
                f"Sheet {sheet_key} row {row_idx}: invalid id {id_val!r} for {nom_val!r} — skipped"
            )
            continue
        if row_id is None:
            parse_errors.append(
                f"Sheet {sheet_key} row {row_idx}: missing id for {nom_val!r} — skipped"
            )
            continue

        # Suggested BC (A/B only)
        suggested_bc: str | None = None
        if col_bc is not None:
            bc_raw = clean(ws.cell(row_idx, col_bc + 1).value)
            if is_bc_number(bc_raw):
                suggested_bc = bc_raw

        parsed = parse_decision(sheet_key, dec_val, suggested_bc)

        rows_out.append({
            "sheet":        sheet_key,
            "ref":          ref_val,
            "id":           row_id,
            "name":         nom_val,
            "decision":     parsed["decision"],
            "suggested_bc": suggested_bc,
            "target_bc":    parsed["target_bc"],
        })

    return rows_out


def main():
    ap = argparse.ArgumentParser(description="Convert merge-decisions xlsx → JSON")
    ap.add_argument("input",  help="Path to merge-clients-review.xlsx")
    ap.add_argument("output", help="Output JSON path")
    args = ap.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ERROR: input file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    parse_errors: list[str] = []
    all_decisions: list[dict] = []

    for sheet_name, sheet_key in SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            parse_errors.append(f"Sheet not found: {sheet_name!r} — skipped")
            continue
        ws = wb[sheet_name]
        rows = parse_sheet(ws, sheet_key, parse_errors)
        all_decisions.extend(rows)
        print(
            f"  Sheet {sheet_key} ({sheet_name}): {len(rows)} rows parsed",
            file=sys.stderr,
        )

    wb.close()

    # Check for duplicate ids across sheets (error — caller should abort)
    seen_ids: dict[int, str] = {}
    dup_errors: list[str] = []
    for d in all_decisions:
        rid = d["id"]
        if rid in seen_ids:
            dup_errors.append(
                f"id={rid} appears in both sheet {seen_ids[rid]} and sheet {d['sheet']} — ABORT"
            )
        else:
            seen_ids[rid] = d["sheet"]
    parse_errors.extend(dup_errors)

    out = {
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "row_count":   len(all_decisions),
        "errors":      parse_errors,
        "decisions":   all_decisions,
    }

    output_path = Path(args.output)
    output_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")

    print(
        f"\nWrote {len(all_decisions)} decisions to {output_path}\n"
        f"Parse errors: {len(parse_errors)}",
        file=sys.stderr,
    )
    if dup_errors:
        print("ABORT: duplicate ids found — fix before sending to PHP script", file=sys.stderr)
        sys.exit(2)


if __name__ == "__main__":
    main()
